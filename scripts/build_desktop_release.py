from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DIST_DIR = ROOT / "dist" / "run_app"
MANIFEST_PATH = ROOT / "release_manifest.json"

DEFAULT_SEED_WORKBOOK_NAMES = [
    "T\u00e4tigkeiten_\u00dcberblick.xlsx",
    "Taetigkeiten_Ueberblick.xlsx",
    "T\u00c3\u00a4tigkeiten_\u00c3\u0153berblick.xlsx",
]
USER_DATA_SHEETS = [
    "T\u00e4tigkeiten",
    "Team_T\u00e4tigkeiten",
    "Meilensteine",
    "Benutzerrechte",
    "Projektrollen",
]
DASHBOARD_SHEET = "Dashboard"
REPORT_SHEET = "Einsatzbericht"
REPORT_RESET_CELLS = ["C12", "H9", "K2", "K3"]
REPORT_DETAIL_START_ROW = 17
REPORT_DETAIL_END_ROW = 35
REPORT_DETAIL_COL_START = 1
REPORT_DETAIL_COL_END = 8


def _seed_workbook_path() -> Path:
    for name in DEFAULT_SEED_WORKBOOK_NAMES:
        candidate = ROOT / "data" / name
        if candidate.exists():
            return candidate
    return ROOT / "data" / "Taetigkeiten_Ueberblick.xlsx"


def _git_version() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=ROOT,
            text=True,
        ).strip()
    except Exception:
        return "dev"


def _load_manifest() -> dict:
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


def _platform_asset_name(platform_name: str) -> str:
    return "EinsatzberichtManager-windows.zip" if platform_name == "windows" else "EinsatzberichtManager-macos.zip"


def _installer_name(platform_name: str) -> str:
    return "install_windows.ps1" if platform_name == "windows" else "install_macos.command"


def _installer_source(platform_name: str) -> Path:
    if platform_name == "windows":
        return ROOT / "deployment" / "windows" / "install_windows.ps1"
    return ROOT / "deployment" / "macos" / "install_macos.command"


def _windows_wrapper_source() -> Path:
    return ROOT / "deployment" / "windows" / "install.bat"


def _copy_runtime_compat_files(app_dir: Path) -> None:
    internal_dir = app_dir / "_internal"
    if not internal_dir.exists():
        return

    for relative_path in [
        Path("release_manifest.json"),
        Path("streamlit_einsatzbericht_app_v2_excel_masterdata.py"),
    ]:
        source = internal_dir / relative_path
        if source.exists():
            target = app_dir / relative_path
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)


def _clear_sheet_values(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    min_row: int = 2,
    min_col: int = 1,
    max_col: int | None = None,
) -> None:
    max_col = max_col or ws.max_column
    if ws.max_row < min_row:
        return
    for row_idx in range(min_row, ws.max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            ws.cell(row_idx, col_idx).value = None


def _reset_report_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for cell_ref in REPORT_RESET_CELLS:
        ws[cell_ref].value = None
    for row_idx in range(REPORT_DETAIL_START_ROW, REPORT_DETAIL_END_ROW + 1):
        for col_idx in range(REPORT_DETAIL_COL_START, REPORT_DETAIL_COL_END + 1):
            ws.cell(row_idx, col_idx).value = None


def _sanitize_distribution_workbook(source: Path, target: Path) -> None:
    wb = openpyxl.load_workbook(source)

    for sheet_name in USER_DATA_SHEETS:
        if sheet_name in wb.sheetnames:
            _clear_sheet_values(wb[sheet_name], min_row=2)

    if DASHBOARD_SHEET in wb.sheetnames:
        _clear_sheet_values(wb[DASHBOARD_SHEET], min_row=1)

    if REPORT_SHEET in wb.sheetnames:
        _reset_report_sheet(wb[REPORT_SHEET])

    calc_props = getattr(wb, "calculation", None)
    if calc_props is not None:
        calc_props.calcMode = "auto"
        calc_props.fullCalcOnLoad = True
        calc_props.forceFullCalc = True

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)


def _seed_workbook_target_paths(app_dir: Path, seed_workbook: Path) -> list[Path]:
    targets: list[Path] = []
    candidate_names = list(dict.fromkeys([seed_workbook.name, *DEFAULT_SEED_WORKBOOK_NAMES]))

    for data_dir in [app_dir / "data", app_dir / "_internal" / "data"]:
        if not data_dir.exists():
            continue
        matched = False
        for name in candidate_names:
            candidate = data_dir / name
            if candidate.exists():
                targets.append(candidate)
                matched = True
        if not matched:
            targets.append(data_dir / seed_workbook.name)

    return list(dict.fromkeys(targets))


def _write_sanitized_seed_workbooks(app_dir: Path, seed_workbook: Path) -> None:
    for target in _seed_workbook_target_paths(app_dir, seed_workbook):
        _sanitize_distribution_workbook(seed_workbook, target)


def _reset_distribution_data_dirs(app_dir: Path) -> None:
    for data_dir in [app_dir / "data", app_dir / "_internal" / "data"]:
        if data_dir.exists():
            shutil.rmtree(data_dir)
        data_dir.mkdir(parents=True, exist_ok=True)


def build_release(platform_name: str, version: str, payload_dir: Path, output_dir: Path) -> Path:
    seed_workbook = _seed_workbook_path()
    if not payload_dir.exists():
        raise FileNotFoundError(f"Build payload not found: {payload_dir}")
    if not seed_workbook.exists():
        raise FileNotFoundError(f"Seed workbook not found: {seed_workbook}")

    staging_dir = output_dir / f"release_{platform_name}"
    app_dir = staging_dir / "app"
    if staging_dir.exists():
        shutil.rmtree(staging_dir)
    app_dir.mkdir(parents=True, exist_ok=True)

    shutil.copytree(payload_dir, app_dir, dirs_exist_ok=True)
    _copy_runtime_compat_files(app_dir)
    _reset_distribution_data_dirs(app_dir)
    _write_sanitized_seed_workbooks(app_dir, seed_workbook)

    manifest = _load_manifest()
    manifest["version"] = version
    manifest["release_asset_windows"] = _platform_asset_name("windows")
    manifest["release_asset_macos"] = _platform_asset_name("macos")
    (app_dir / "release_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    installer_source = _installer_source(platform_name)
    shutil.copy2(installer_source, staging_dir / _installer_name(platform_name))
    if platform_name == "windows":
        shutil.copy2(_windows_wrapper_source(), staging_dir / "install.bat")

    readme = staging_dir / "README_INSTALL.txt"
    starter_name = "install.bat" if platform_name == "windows" else _installer_name(platform_name)
    bundled_runtime = (payload_dir / "run_app.exe").exists()
    dependency_note = (
        "6. Im gebuendelten Tester-ZIP sind die benoetigten Python-Pakete bereits enthalten."
        if bundled_runtime
        else "6. Dieses Source-Installer-ZIP erstellt beim Installieren eine lokale Python-Umgebung und installiert die benoetigten Pakete."
    )
    readme.write_text(
        "\n".join(
            [
                "Einsatzbericht Manager Testdistribution",
                "",
                f"Version: {version}",
                "",
                "1. Entpacke diese ZIP-Datei.",
                f"2. Starte {starter_name}.",
                "3. Die App kopiert sich in dein Benutzerprofil und legt eine Startverknuepfung an.",
                "4. Updates werden beim Start ueber GitHub Releases geprueft.",
                "5. Die mitgelieferte Excel-Datei ist eine leere Startvorlage ohne Nutzerdaten.",
                dependency_note,
                "",
                "Hinweis:",
                "Wenn das Repository privat bleibt, muessen die Release-ZIPs oeffentlich oder ueber einen anderen Download-Kanal erreichbar sein.",
            ]
        ),
        encoding="utf-8",
    )

    zip_path = output_dir / _platform_asset_name(platform_name)
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in staging_dir.rglob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(staging_dir))
    return zip_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a tester ZIP for the desktop app.")
    parser.add_argument("--platform", choices=["windows", "macos"], required=True)
    parser.add_argument("--version", default=_git_version())
    parser.add_argument("--payload-dir", default=str(DIST_DIR))
    parser.add_argument("--output-dir", default=str(ROOT / "release"))
    args = parser.parse_args()

    zip_path = build_release(
        platform_name=args.platform,
        version=args.version,
        payload_dir=Path(args.payload_dir).resolve(),
        output_dir=Path(args.output_dir).resolve(),
    )
    print(f"Created release ZIP: {zip_path}")


if __name__ == "__main__":
    main()

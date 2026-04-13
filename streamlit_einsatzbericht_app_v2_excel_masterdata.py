from __future__ import annotations

import datetime as dt
import math
import os
import shutil
import sys
import subprocess
import tempfile
import json
import re
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

try:
    import streamlit as st
except ModuleNotFoundError:  # pragma: no cover
    st = None

APP_TITLE = "Einsatzbericht Manager (MVP)"
TAETIGKEITEN_SHEET = "Tätigkeiten"
TEAM_SHEET = "Team_Tätigkeiten"
HILFS_SHEET = "Hilfstabelle"
KODIERUNG_SHEET = "Kodierung Joyson"
RELEVANTE_KODIERUNG_SHEET = "relevante Kodierung"

# Original Excel report template area (detail rows)
REPORT_DETAIL_START_ROW = 17
REPORT_DETAIL_END_ROW = 35
REPORT_DETAIL_COL_START = 1  # A
REPORT_DETAIL_COL_END = 8  # H
REPORT_ROWS_PER_PAGE = REPORT_DETAIL_END_ROW - REPORT_DETAIL_START_ROW + 1
MILESTONES_SHEET = "Meilensteine"
MILESTONE_COLS = ["Projekt", "Meilenstein", "Datum", "Status", "Fortschritt", "Kommentar"]
MILESTONE_STATUSES = ["geplant", "in arbeit", "blockiert", "erledigt"]

TAET_COLS = [
    "Datum",
    "Projekt",
    "Zeit von",
    "Zeit bis",
    "Pause",
    "Stunden",
    "Zahl",
    "km",
    "Tätigkeit",
    "Kodierung",
    "Interne Projekte",
    "Info",
    "Abgerechnet",
    "eingetragen",
]

TEAM_COLS = ["Mitarbeiter"] + TAET_COLS

KEY_COLS_FOR_EMPTY_CHECK = [1, 2, 3, 4, 9, 12, 13, 14]  # ignore formula columns F/G
TEAM_KEY_COLS_FOR_EMPTY_CHECK = [2, 3, 4, 5, 10, 13, 14, 15]


@dataclass
class WorkbookData:
    path: Path
    taetigkeiten_df: pd.DataFrame
    team_df: pd.DataFrame
    lookups: Dict[str, Any]
    milestones_df: pd.DataFrame


# ------------------------- parsing helpers -------------------------

def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_yes_no(value: Any) -> str:
    s = _safe_str(value).strip().lower()
    if s in {"ja", "yes", "y", "true", "1"}:
        return "ja"
    if s in {"nein", "no", "n", "false", "0"}:
        return "nein"
    return ""


def _to_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def _to_time(value: Any) -> Optional[dt.time]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, dt.time):
        return value.replace(microsecond=0)
    if isinstance(value, dt.timedelta):
        total_seconds = int(value.total_seconds())
        total_seconds %= 24 * 3600
        return dt.time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
    if isinstance(value, (int, float)) and not pd.isna(value):
        frac = float(value) % 1
        total_seconds = int(round(frac * 24 * 3600))
        total_seconds %= 24 * 3600
        return dt.time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return dt.datetime.strptime(s, fmt).time()
            except ValueError:
                pass
    return None


def _time_to_minutes(value: Any) -> int:
    t = _to_time(value)
    if t is None:
        return 0
    return t.hour * 60 + t.minute + round(t.second / 60)


def _minutes_to_time(minutes: int) -> Optional[dt.time]:
    if minutes <= 0:
        return None
    minutes = int(minutes)
    minutes = max(0, minutes)
    h = (minutes // 60) % 24
    m = minutes % 60
    return dt.time(hour=h, minute=m)


def _compute_hours_decimal(start: Any, end: Any, pause_minutes: int) -> Optional[float]:
    t_start = _to_time(start)
    t_end = _to_time(end)
    if t_start is None or t_end is None:
        return None
    d0 = dt.datetime.combine(dt.date(2000, 1, 1), t_start)
    d1 = dt.datetime.combine(dt.date(2000, 1, 1), t_end)
    if d1 < d0:
        d1 += dt.timedelta(days=1)
    delta = d1 - d0 - dt.timedelta(minutes=max(pause_minutes, 0))
    hours = delta.total_seconds() / 3600
    if hours < 0:
        return None
    return round(hours, 4)


def _format_time(value: Any) -> str:
    t = _to_time(value)
    if t is None:
        return ""
    return t.strftime("%H:%M")


def _format_date(value: Any) -> str:
    d = _to_date(value)
    if d is None:
        return ""
    return d.strftime("%d.%m.%Y")


def _display_row_label(row: pd.Series) -> str:
    d = _format_date(row.get("Datum"))
    zv = _format_time(row.get("Zeit von"))
    zb = _format_time(row.get("Zeit bis"))
    projekt = _safe_str(row.get("Projekt"))
    typ = _safe_str(row.get("Tätigkeit"))
    info = _safe_str(row.get("Info"))[:40]
    return f"Zeile {int(row['_excel_row'])} | {d} | {projekt} | {zv}-{zb} | {typ} | {info}"


# ------------------------- workbook reading -------------------------

def _read_list_column(ws, col_idx: int, start_row: int = 2) -> List[str]:
    values: List[str] = []
    seen = set()
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if _is_blank(v):
            continue
        s = _safe_str(v).strip()
        if s not in seen:
            seen.add(s)
            values.append(s)
    return values


def _load_lookups(wb: openpyxl.Workbook) -> Dict[str, Any]:
    lookups: Dict[str, Any] = {}

    if HILFS_SHEET not in wb.sheetnames:
        raise ValueError(f"Arbeitsblatt '{HILFS_SHEET}' nicht gefunden.")
    h = wb[HILFS_SHEET]

    lookups["ja_nein"] = _read_list_column(h, 1)
    if not lookups["ja_nein"]:
        lookups["ja_nein"] = ["ja", "nein"]

    lookups["taetigkeit_typen"] = _read_list_column(h, 3)
    if not lookups["taetigkeit_typen"]:
        lookups["taetigkeit_typen"] = ["F", "R", "I", "S", "K"]

    lookups["interne_projekte"] = _read_list_column(h, 5)
    lookups["projekte"] = _read_list_column(h, 7)

    projekt_infos: Dict[str, Dict[str, Any]] = {}
    for r in range(2, h.max_row + 1):
        projekt = h.cell(r, 7).value
        if _is_blank(projekt):
            continue
        p = _safe_str(projekt).strip()
        projekt_infos[p] = {
            "Projekt": p,
            "Kunde": h.cell(r, 8).value,
            "Straße": h.cell(r, 9).value,
            "Ort": h.cell(r, 10).value,
            "Ansprechpartner": h.cell(r, 11).value,
            "Projektadresse Standard": h.cell(r, 12).value,
            "Projektadresse Alternativ": h.cell(r, 13).value,
        }
    lookups["projekt_infos"] = projekt_infos

    kodierung_map_eb: Dict[str, str] = {}
    kodierung_map_intern: Dict[str, str] = {}
    kodierungen_aufgaben: List[str] = []
    if KODIERUNG_SHEET in wb.sheetnames:
        k = wb[KODIERUNG_SHEET]
        seen = set()
        for r in range(2, k.max_row + 1):
            aufgabe = k.cell(r, 2).value  # B
            if _is_blank(aufgabe):
                continue
            aufgabe_s = _safe_str(aufgabe).strip()
            if aufgabe_s not in seen:
                kodierungen_aufgaben.append(aufgabe_s)
                seen.add(aufgabe_s)
            kod_intern = k.cell(r, 3).value  # C
            kod_eb = k.cell(r, 4).value  # D
            if not _is_blank(kod_eb):
                kodierung_map_eb[aufgabe_s] = _safe_str(kod_eb).strip()
            if not _is_blank(kod_intern):
                kodierung_map_intern[aufgabe_s] = _safe_str(kod_intern).strip()
    lookups["kodierung_aufgaben"] = kodierungen_aufgaben
    lookups["kodierung_map_eb"] = kodierung_map_eb
    lookups["kodierung_map_intern"] = kodierung_map_intern

    relevante: List[str] = []
    if KODIERUNG_SHEET in wb.sheetnames:
        k = wb[KODIERUNG_SHEET]
        for r in range(2, k.max_row + 1):
            marker = _safe_str(k.cell(r, 1).value).strip().lower()
            aufgabe = k.cell(r, 2).value
            if marker == "x" and not _is_blank(aufgabe):
                relevante.append(_safe_str(aufgabe).strip())
    if not relevante and RELEVANTE_KODIERUNG_SHEET in wb.sheetnames:
        rk = wb[RELEVANTE_KODIERUNG_SHEET]
        relevante = _read_list_column(rk, 1)
    lookups["relevante_kodierungen"] = list(dict.fromkeys(relevante))

    return lookups


def _read_taetigkeiten_df(wb: openpyxl.Workbook) -> pd.DataFrame:
    if TAETIGKEITEN_SHEET not in wb.sheetnames:
        raise ValueError(f"Arbeitsblatt '{TAETIGKEITEN_SHEET}' nicht gefunden.")
    ws = wb[TAETIGKEITEN_SHEET]

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, 15)]
        is_empty = all(_is_blank(ws.cell(r, c).value) for c in KEY_COLS_FOR_EMPTY_CHECK)
        if is_empty:
            continue

        rec = dict(zip(TAET_COLS, raw))
        rec["_excel_row"] = r
        rec["Datum"] = _to_date(rec.get("Datum"))
        rec["Zeit von"] = _to_time(rec.get("Zeit von"))
        rec["Zeit bis"] = _to_time(rec.get("Zeit bis"))
        rec["Pause"] = _to_time(rec.get("Pause"))
        rec["Pause_Min"] = _time_to_minutes(rec.get("Pause"))

        zahl = rec.get("Zahl")
        if isinstance(zahl, (int, float)) and not (isinstance(zahl, float) and math.isnan(zahl)):
            rec["Zahl"] = float(zahl)
        else:
            rec["Zahl"] = _compute_hours_decimal(rec.get("Zeit von"), rec.get("Zeit bis"), rec.get("Pause_Min", 0))

        rec["Stunden_Anzeige"] = None
        if rec.get("Zahl") is not None:
            total_minutes = int(round(float(rec["Zahl"]) * 60))
            rec["Stunden_Anzeige"] = f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"

        rec["Abgerechnet"] = _normalize_yes_no(rec.get("Abgerechnet")) or _safe_str(rec.get("Abgerechnet"))
        rec["eingetragen"] = _normalize_yes_no(rec.get("eingetragen")) or _safe_str(rec.get("eingetragen"))

        rows.append(rec)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=TAET_COLS + ["_excel_row", "Pause_Min", "Stunden_Anzeige"])
        return df

    sort_date = pd.to_datetime(df["Datum"], errors="coerce")
    sort_start = df["Zeit von"].apply(lambda x: _time_to_minutes(x) if x else -1)
    df = df.assign(_sort_date=sort_date, _sort_start=sort_start).sort_values(
        ["_sort_date", "_sort_start", "_excel_row"], ascending=[False, False, False]
    )
    df = df.drop(columns=["_sort_date", "_sort_start"]).reset_index(drop=True)
    return df


def _read_team_df(wb: openpyxl.Workbook) -> pd.DataFrame:
    if TEAM_SHEET not in wb.sheetnames:
        return pd.DataFrame(columns=TEAM_COLS + ["_excel_row", "Pause_Min", "Stunden_Anzeige"])

    ws = wb[TEAM_SHEET]
    rows: List[Dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, 16)]
        is_empty = all(_is_blank(ws.cell(r, c).value) for c in TEAM_KEY_COLS_FOR_EMPTY_CHECK)
        if is_empty:
            continue

        rec = dict(zip(TEAM_COLS, raw))
        rec["_excel_row"] = r
        rec["Datum"] = _to_date(rec.get("Datum"))
        rec["Zeit von"] = _to_time(rec.get("Zeit von"))
        rec["Zeit bis"] = _to_time(rec.get("Zeit bis"))
        rec["Pause"] = _to_time(rec.get("Pause"))
        rec["Pause_Min"] = _time_to_minutes(rec.get("Pause"))
        rec["Mitarbeiter"] = _safe_str(rec.get("Mitarbeiter")).strip()

        zahl = rec.get("Zahl")
        if isinstance(zahl, (int, float)) and not (isinstance(zahl, float) and math.isnan(zahl)):
            rec["Zahl"] = float(zahl)
        else:
            rec["Zahl"] = _compute_hours_decimal(rec.get("Zeit von"), rec.get("Zeit bis"), rec.get("Pause_Min", 0))

        rec["Stunden_Anzeige"] = None
        if rec.get("Zahl") is not None:
            total_minutes = int(round(float(rec["Zahl"]) * 60))
            rec["Stunden_Anzeige"] = f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"

        rec["Abgerechnet"] = _normalize_yes_no(rec.get("Abgerechnet")) or _safe_str(rec.get("Abgerechnet"))
        rec["eingetragen"] = _normalize_yes_no(rec.get("eingetragen")) or _safe_str(rec.get("eingetragen"))

        rows.append(rec)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=TEAM_COLS + ["_excel_row", "Pause_Min", "Stunden_Anzeige"])
        return df

    sort_date = pd.to_datetime(df["Datum"], errors="coerce")
    sort_start = df["Zeit von"].apply(lambda x: _time_to_minutes(x) if x else -1)
    df = df.assign(_sort_date=sort_date, _sort_start=sort_start).sort_values(
        ["_sort_date", "_sort_start", "_excel_row"], ascending=[False, False, False]
    )
    df = df.drop(columns=["_sort_date", "_sort_start"]).reset_index(drop=True)
    return df


def _default_excel_candidates() -> List[Path]:
    script_dir = Path(__file__).resolve().parent
    cwd = Path.cwd()
    names = [
        "Tätigkeiten_Überblick.xlsx",
        "__Tätigkeiten_Überblick - Kopie.xlsx",
    ]
    cands: List[Path] = []
    for base in [script_dir, script_dir / "data", cwd, cwd / "data"]:
        for n in names:
            cands.append(base / n)
        for p in sorted(base.glob("*.xlsx")):
            cands.append(p)
    cands.append(Path("/mnt/data/__Tätigkeiten_Überblick - Kopie.xlsx"))
    out: List[Path] = []
    seen = set()
    for c in cands:
        key = str(c)
        if key not in seen:
            seen.add(key)
            out.append(c)
    return out


def _resolve_excel_path(path_str: str) -> Path:
    raw = (path_str or "").strip()
    if raw:
        p = Path(raw).expanduser()
        try:
            if p.exists():
                return p.resolve()
        except Exception:
            pass
        candidates = [
            Path(__file__).resolve().parent / p,
            Path.cwd() / p,
        ]
        for c in candidates:
            if c.exists():
                return c.resolve()
        return p.resolve()
    for c in _default_excel_candidates():
        if c.exists():
            return c.resolve()
    return (Path(__file__).resolve().parent / "data" / "Tätigkeiten_Überblick.xlsx").resolve()


def _store_uploaded_excel(uploaded_file) -> Path:
    if uploaded_file is None:
        raise ValueError("Keine Datei hochgeladen.")

    original_name = Path(getattr(uploaded_file, "name", "upload.xlsx")).name
    suffix = Path(original_name).suffix.lower()

    if suffix != ".xlsx":
        raise ValueError("Bitte eine .xlsx-Datei hochladen.")

    script_dir = Path(__file__).resolve().parent
    imports_dir = script_dir / "imports"
    imports_dir.mkdir(parents=True, exist_ok=True)

    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    stem = Path(original_name).stem
    target = imports_dir / f"{stem}_import_{ts}.xlsx"

    data = uploaded_file.getvalue()
    target.write_bytes(data)

    return target.resolve()


# ------------------------- Cross-Platform Excel Automation -------------------------

def _prepare_report_formula_in_excel_sheet_com(ws) -> None:
    try:
        ws.Range(f"A{REPORT_DETAIL_START_ROW}:H200").ClearContents()
    except Exception:
        ws.Range(f"A{REPORT_DETAIL_START_ROW}:H{REPORT_DETAIL_END_ROW}").ClearContents()
    formula_en = '=FILTER(Berechnung!C2:J444,Berechnung!B2:B444=Einsatzbericht!C12,"")'
    try:
        ws.Range(f"A{REPORT_DETAIL_START_ROW}").Formula2 = formula_en
    except Exception:
        ws.Range(
            f"A{REPORT_DETAIL_START_ROW}").FormulaLocal = '=FILTER(Berechnung!C2:J444;Berechnung!B2:B444=Einsatzbericht!C12;"")'


def _clear_original_report_detail_area_com(ws) -> None:
    ws.Range(f"A{REPORT_DETAIL_START_ROW}:H{REPORT_DETAIL_END_ROW}").ClearContents()


def _write_original_report_page_com(ws, page_df: pd.DataFrame, year: int, month: int, project: str) -> None:
    ws.Range("K2").Value = int(year)
    ws.Range("K3").Value = int(month)
    ws.Range("C12").Value = str(project)

    _clear_original_report_detail_area_com(ws)

    if page_df is None or page_df.empty:
        return

    page_df = page_df.reset_index(drop=True)
    max_rows = min(len(page_df), REPORT_ROWS_PER_PAGE)
    rows_2d = []
    for idx in range(max_rows):
        rows_2d.append(_report_row_to_excel_values(page_df.iloc[idx]))

    if rows_2d:
        start = REPORT_DETAIL_START_ROW
        end = REPORT_DETAIL_START_ROW + len(rows_2d) - 1
        ws.Range(f"A{start}:H{end}").Value = tuple(tuple(r) for r in rows_2d)


def _prepare_report_formula_in_excel_sheet_openpyxl(ws) -> None:
    for r in range(REPORT_DETAIL_START_ROW, REPORT_DETAIL_END_ROW + 1):
        for c in range(REPORT_DETAIL_COL_START, REPORT_DETAIL_COL_END + 1):
            ws.cell(r, c).value = None
    formula_en = '=FILTER(Berechnung!C2:J444,Berechnung!B2:B444=Einsatzbericht!C12,"")'
    ws.cell(REPORT_DETAIL_START_ROW, 1).value = formula_en


def _write_original_report_page_openpyxl(ws, page_df: pd.DataFrame, year: int, month: int, project: str) -> None:
    ws["K2"].value = int(year)
    ws["K3"].value = int(month)
    ws["C12"].value = str(project)

    for r in range(REPORT_DETAIL_START_ROW, REPORT_DETAIL_END_ROW + 1):
        for c in range(REPORT_DETAIL_COL_START, REPORT_DETAIL_COL_END + 1):
            ws.cell(r, c).value = None

    if page_df is None or page_df.empty:
        return

    page_df = page_df.reset_index(drop=True)
    max_rows = min(len(page_df), REPORT_ROWS_PER_PAGE)
    rows_2d = []
    for idx in range(max_rows):
        rows_2d.append(_report_row_to_excel_values(page_df.iloc[idx]))

    if rows_2d:
        start = REPORT_DETAIL_START_ROW
        for row_offset, row_data in enumerate(rows_2d):
            for col_offset, val in enumerate(row_data):
                ws.cell(start + row_offset, REPORT_DETAIL_COL_START + col_offset).value = val


def _report_row_to_excel_values(row: pd.Series) -> List[Any]:
    datum = _format_date(row.get("Datum")) or _safe_str(row.get("Datum"))
    beginn = _format_time(row.get("Beginn")) or _safe_str(row.get("Beginn"))
    ende = _format_time(row.get("Ende")) or _safe_str(row.get("Ende"))
    pause = _format_time(row.get("Pause")) or _safe_str(row.get("Pause"))

    zeit_h = row.get("Zeit (h)")
    if zeit_h is None or (isinstance(zeit_h, float) and math.isnan(zeit_h)):
        zeit_h = None
    else:
        try:
            zeit_h = float(zeit_h)
        except Exception:
            zeit_h = None

    art = _safe_str(row.get("Art"))
    kod_eb = row.get("Kodierung EB")
    if pd.isna(kod_eb):
        kod_eb = None
    else:
        kod_eb = _safe_str(kod_eb) or None
    leistung = _safe_str(row.get("Leistungsbeschreibung"))
    return [datum, beginn, ende, pause, zeit_h, art, kod_eb, leistung]


def _split_report_pages(report_df: Optional[pd.DataFrame]) -> List[pd.DataFrame]:
    if report_df is None:
        return []
    if report_df.empty:
        return [report_df.copy()]
    pages: List[pd.DataFrame] = []
    total = len(report_df)
    for start in range(0, total, REPORT_ROWS_PER_PAGE):
        pages.append(report_df.iloc[start:start + REPORT_ROWS_PER_PAGE].copy().reset_index(drop=True))
    return pages


def _excel_original_report_action_com(
        xlsx_path: Path,
        year: int,
        month: int,
        project: str,
        action: str,
        pdf_output_path: Optional[Path] = None,
        xlsx_output_path: Optional[Path] = None,
        report_df: Optional[pd.DataFrame] = None,
) -> Tuple[bool, str, List[Path]]:
    """Original Windows-only COM automation path via pywin32."""
    try:
        import pythoncom  # type: ignore
        import win32com.client as win32  # type: ignore
    except ImportError:
        raise RuntimeError("win32com is not available.")

    if not Path(xlsx_path).exists():
        return False, f"Excel-Datei nicht gefunden: {xlsx_path}", []

    excel = None
    wb = None
    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = bool(action == "open")
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        try:
            excel.EnableEvents = False
        except Exception:
            pass

        manual_calc_enabled = False
        try:
            excel.Calculation = -4135  # xlCalculationManual
            manual_calc_enabled = True
        except Exception:
            manual_calc_enabled = False

        wb = excel.Workbooks.Open(str(Path(xlsx_path).resolve()), UpdateLinks=False, ReadOnly=False)
        ws = wb.Worksheets("Einsatzbericht")

        pages = _split_report_pages(report_df)
        if not pages:
            pages = [pd.DataFrame()]
        page_count = len(pages)

        def _recalc() -> None:
            try:
                ws.Calculate()
                return
            except Exception:
                pass
            try:
                excel.Calculate()
                return
            except Exception:
                pass
            try:
                excel.CalculateFull()
            except Exception:
                pass

        if action == "pdf":
            if pdf_output_path is None:
                pdf_output_path = Path(xlsx_path).with_name(f"Einsatzbericht_{project}_{year}-{int(month):02d}.pdf")
            pdf_output_path = Path(pdf_output_path)
            pdf_output_path.parent.mkdir(parents=True, exist_ok=True)

            exported_files: List[Path] = []
            for page_idx, page_df in enumerate(pages, start=1):
                if report_df is not None:
                    _write_original_report_page_com(ws, page_df, int(year), int(month), str(project))
                else:
                    ws.Range("K2").Value = int(year)
                    ws.Range("K3").Value = int(month)
                    ws.Range("C12").Value = str(project)
                    _prepare_report_formula_in_excel_sheet_com(ws)
                _recalc()

                if page_count > 1:
                    out_path = pdf_output_path.with_name(
                        f"{pdf_output_path.stem}_{page_idx:02d}{pdf_output_path.suffix}")
                else:
                    out_path = pdf_output_path
                ws.ExportAsFixedFormat(0, str(out_path))
                exported_files.append(out_path)

            wb.Close(SaveChanges=False)
            try:
                if manual_calc_enabled:
                    excel.Calculation = -4105  # xlCalculationAutomatic
            except Exception:
                pass
            excel.Quit()
            if page_count > 1:
                return True, f"{page_count} PDFs exportiert ({REPORT_ROWS_PER_PAGE} Positionen pro Formularseite).", exported_files
            return True, f"PDF exportiert.", exported_files

        if action == "xlsx":
            if xlsx_output_path is None:
                xlsx_output_path = Path(xlsx_path).with_name(f"Einsatzbericht_{project}_{year}-{int(month):02d}.xlsx")
            xlsx_output_path = Path(xlsx_output_path)
            xlsx_output_path.parent.mkdir(parents=True, exist_ok=True)

            exported_files: List[Path] = []
            for page_idx, page_df in enumerate(pages, start=1):
                if report_df is not None:
                    _write_original_report_page_com(ws, page_df, int(year), int(month), str(project))
                else:
                    ws.Range("K2").Value = int(year)
                    ws.Range("K3").Value = int(month)
                    ws.Range("C12").Value = str(project)
                    _prepare_report_formula_in_excel_sheet_com(ws)
                _recalc()

                if page_count > 1:
                    out_path = xlsx_output_path.with_name(
                        f"{xlsx_output_path.stem}_{page_idx:02d}{xlsx_output_path.suffix}")
                else:
                    out_path = xlsx_output_path

                wb.SaveCopyAs(str(out_path))
                exported_files.append(out_path)

            wb.Close(SaveChanges=False)
            try:
                if manual_calc_enabled:
                    excel.Calculation = -4105
            except Exception:
                pass
            excel.Quit()
            if page_count > 1:
                return True, f"{page_count} Excel-Kopien exportiert ({REPORT_ROWS_PER_PAGE} Positionen pro Formularseite).", exported_files
            return True, f"Excel-Kopie exportiert.", exported_files

        if action == "print":
            for page_df in pages:
                if report_df is not None:
                    _write_original_report_page_com(ws, page_df, int(year), int(month), str(project))
                else:
                    ws.Range("K2").Value = int(year)
                    ws.Range("K3").Value = int(month)
                    ws.Range("C12").Value = str(project)
                    _prepare_report_formula_in_excel_sheet_com(ws)
                _recalc()
                ws.PrintOut()
            wb.Close(SaveChanges=False)
            try:
                if manual_calc_enabled:
                    excel.Calculation = -4105
            except Exception:
                pass
            excel.Quit()
            if page_count > 1:
                return True, f"{page_count} Druckaufträge gesendet ({REPORT_ROWS_PER_PAGE} Positionen pro Formularseite).", []
            return True, "Druckauftrag an Standarddrucker gesendet.", []

        if action == "open":
            if report_df is not None:
                _write_original_report_page_com(ws, pages[0], int(year), int(month), str(project))
            else:
                ws.Range("K2").Value = int(year)
                ws.Range("K3").Value = int(month)
                ws.Range("C12").Value = str(project)
                _prepare_report_formula_in_excel_sheet_com(ws)
            _recalc()
            excel.Visible = True
            excel.ScreenUpdating = True
            excel.DisplayAlerts = True
            if page_count > 1:
                return True, f"Originaldatei in Excel geöffnet (Seite 1/{page_count} vorbereitet). PDF/Druck erstellt weitere Seiten automatisch.", []
            return True, "Originaldatei in Excel geöffnet (Bericht vorbereitet).", []

        if wb is not None:
            wb.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        return False, f"Unbekannte Aktion: {action}", []

    except Exception as e:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        raise e
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _excel_original_report_action_fallback(
        xlsx_path: Path,
        year: int,
        month: int,
        project: str,
        action: str,
        pdf_output_path: Optional[Path] = None,
        xlsx_output_path: Optional[Path] = None,
        report_df: Optional[pd.DataFrame] = None,
        is_mac: bool = False,
) -> Tuple[bool, str, List[Path]]:
    """Cross-Platform Fallback (macOS / Linux / fallback for Windows without win32com)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        if "Einsatzbericht" not in wb.sheetnames:
            return False, "Blatt 'Einsatzbericht' nicht gefunden.", []
        ws = wb["Einsatzbericht"]

        pages = _split_report_pages(report_df)
        if not pages:
            pages = [pd.DataFrame()]
        page_count = len(pages)

        exported_files: List[Path] = []

        if action == "pdf":
            if pdf_output_path is None:
                pdf_output_path = Path(xlsx_path).with_name(f"Einsatzbericht_{project}_{year}-{int(month):02d}.pdf")
            pdf_output_path = Path(pdf_output_path)
            pdf_output_path.parent.mkdir(parents=True, exist_ok=True)
        elif action == "xlsx":
            if xlsx_output_path is None:
                xlsx_output_path = Path(xlsx_path).with_name(f"Einsatzbericht_{project}_{year}-{int(month):02d}.xlsx")
            xlsx_output_path = Path(xlsx_output_path)
            xlsx_output_path.parent.mkdir(parents=True, exist_ok=True)

        for page_idx, page_df in enumerate(pages, start=1):
            if report_df is not None:
                _write_original_report_page_openpyxl(ws, page_df, year, month, project)
            else:
                ws["K2"].value = int(year)
                ws["K3"].value = int(month)
                ws["C12"].value = str(project)
                _prepare_report_formula_in_excel_sheet_openpyxl(ws)

            if action == "xlsx":
                if page_count > 1:
                    out_path = xlsx_output_path.with_name(
                        f"{xlsx_output_path.stem}_{page_idx:02d}{xlsx_output_path.suffix}")
                else:
                    out_path = xlsx_output_path
                wb.save(out_path)
                exported_files.append(out_path)

            elif action == "open":
                tmp_dir = Path(tempfile.gettempdir())
                tmp_file = tmp_dir / f"Einsatzbericht_Temp_{project}_{year}-{int(month):02d}_{page_idx:02d}.xlsx"
                wb.save(tmp_file)
                if is_mac:
                    subprocess.call(["open", str(tmp_file)])
                elif sys.platform == "win32":
                    os.startfile(str(tmp_file))
                else:
                    subprocess.call(["xdg-open", str(tmp_file)])
                exported_files.append(tmp_file)

            elif action == "pdf":
                if not is_mac:
                    return False, "Direkter PDF-Export erfordert Windows (pywin32) oder macOS (AppleScript).", []

                if page_count > 1:
                    out_path = pdf_output_path.with_name(
                        f"{pdf_output_path.stem}_{page_idx:02d}{pdf_output_path.suffix}")
                else:
                    out_path = pdf_output_path

                tmp_file = Path(tempfile.gettempdir()) / f"tmp_pdf_export_{page_idx}.xlsx"
                wb.save(tmp_file)

                script = f'''
                tell application "Microsoft Excel"
                    open (POSIX file "{tmp_file.resolve()}")
                    save active workbook in (POSIX file "{out_path.resolve()}") as PDF file format
                    close active workbook saving no
                end tell
                '''
                subprocess.run(["osascript", "-e", script], check=True)
                exported_files.append(out_path)

            elif action == "print":
                if not is_mac:
                    return False, "Direkter Druck erfordert Windows (pywin32) oder macOS (AppleScript).", []
                tmp_file = Path(tempfile.gettempdir()) / f"tmp_print_export_{page_idx}.xlsx"
                wb.save(tmp_file)
                script = f'''
                tell application "Microsoft Excel"
                    open (POSIX file "{tmp_file.resolve()}")
                    print active sheet
                    close active workbook saving no
                end tell
                '''
                subprocess.run(["osascript", "-e", script], check=True)

        if action == "pdf":
            msg = f"{page_count} PDFs exportiert" if page_count > 1 else "PDF exportiert"
            return True, f"{msg}", exported_files
        elif action == "xlsx":
            msg = f"{page_count} Excel-Kopien exportiert" if page_count > 1 else "Excel-Kopie exportiert"
            return True, f"{msg}", exported_files
        elif action == "open":
            return True, f"Datei(en) geöffnet ({len(exported_files)} Seite(n) vorbereitet).", exported_files
        elif action == "print":
            return True, f"{page_count} Druckaufträge via AppleScript an Excel gesendet.", []

        return False, f"Unbekannte Aktion: {action}", []

    except Exception as e:
        return False, f"Fehler bei macOS/Cross-Platform Verarbeitung: {e}", []


def _excel_original_report_action(
        xlsx_path: Path,
        year: int,
        month: int,
        project: str,
        action: str,
        pdf_output_path: Optional[Path] = None,
        xlsx_output_path: Optional[Path] = None,
        report_df: Optional[pd.DataFrame] = None,
) -> Tuple[bool, str, List[Path]]:
    """
    Main dispatcher: Uses COM on Windows if available. Otherwise, falls back
    to a robust cross-platform implementation (openpyxl + osascript on macOS).
    """
    is_windows = sys.platform == "win32"
    is_mac = sys.platform == "darwin"

    if is_windows:
        try:
            return _excel_original_report_action_com(
                xlsx_path, year, month, project, action, pdf_output_path, xlsx_output_path, report_df
            )
        except Exception:
            pass  # Fallthrough to fallback

    return _excel_original_report_action_fallback(
        xlsx_path, year, month, project, action, pdf_output_path, xlsx_output_path, report_df, is_mac
    )


def load_workbook_data(path_str: str) -> WorkbookData:
    path = _resolve_excel_path(path_str)
    if not path.exists():
        raise FileNotFoundError(f"Datei nicht gefunden: {path}")
    wb = openpyxl.load_workbook(path)
    lookups = _load_lookups(wb)
    taetigkeiten_df = _read_taetigkeiten_df(wb)
    team_df = _read_team_df(wb)
    milestones_df = _read_milestones_df(wb)
    return WorkbookData(path=path, taetigkeiten_df=taetigkeiten_df, team_df=team_df, lookups=lookups,
                        milestones_df=milestones_df)


@st.cache_data(show_spinner=False)
def _cached_load_workbook_data(path_str: str, modified_time: float) -> WorkbookData:
    return load_workbook_data(path_str)


# --------------------- Milestone ------------------------------------

def _normalize_milestone_status(v: Any) -> str:
    s = _safe_str(v).strip().lower()
    if not s:
        return "geplant"
    if s in {"done", "erledigt", "fertig"}:
        return "erledigt"
    if s in {"blocked", "blockiert"}:
        return "blockiert"
    if s in {"in progress", "in arbeit", "arbeit"}:
        return "in arbeit"
    if s in {"planned", "geplant"}:
        return "geplant"
    return s


def _read_milestones_df(wb: openpyxl.Workbook) -> pd.DataFrame:
    if MILESTONES_SHEET not in wb.sheetnames:
        return pd.DataFrame(columns=MILESTONE_COLS + ["_excel_row"])

    ws = wb[MILESTONES_SHEET]
    rows: List[Dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        projekt = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        datum = ws.cell(r, 3).value
        status = ws.cell(r, 4).value
        fort = ws.cell(r, 5).value
        comment = ws.cell(r, 6).value

        if _is_blank(projekt) and _is_blank(name) and _is_blank(datum) and _is_blank(status) and _is_blank(
                fort) and _is_blank(comment):
            continue

        d = _to_date(datum)
        f = _to_float_or_none(fort)
        if f is None:
            f = 0.0
        f = max(0.0, min(100.0, float(f)))

        rec = {
            "Projekt": _safe_str(projekt).strip(),
            "Meilenstein": _safe_str(name).strip(),
            "Datum": d,
            "Status": _normalize_milestone_status(status),
            "Fortschritt": f,
            "Kommentar": _safe_str(comment).strip(),
            "_excel_row": r,
        }
        rows.append(rec)

    dfm = pd.DataFrame(rows)
    if dfm.empty:
        return pd.DataFrame(columns=MILESTONE_COLS + ["_excel_row"])

    dfm["Datum_dt"] = pd.to_datetime(dfm["Datum"], errors="coerce")
    dfm = dfm.sort_values(["Projekt", "Datum_dt", "_excel_row"], ascending=[True, True, True]).drop(
        columns=["Datum_dt"])
    dfm = dfm.reset_index(drop=True)
    return dfm


def _ensure_milestones_sheet(wb: openpyxl.Workbook):
    if MILESTONES_SHEET in wb.sheetnames:
        ws = wb[MILESTONES_SHEET]
    else:
        ws = wb.create_sheet(MILESTONES_SHEET)

    if ws.max_row < 1 or all(_is_blank(ws.cell(1, c).value) for c in range(1, 7)):
        for i, h in enumerate(MILESTONE_COLS, start=1):
            ws.cell(1, i).value = h
    return ws


def _find_next_milestone_row(ws) -> int:
    for r in range(2, ws.max_row + 2):
        if all(_is_blank(ws.cell(r, c).value) for c in range(1, 7)):
            return r
    return ws.max_row + 1


def _write_milestone_row(ws, row_idx: int, rec: Dict[str, Any]) -> None:
    projekt = _safe_str(rec.get("Projekt")).strip()
    name = _safe_str(rec.get("Meilenstein")).strip()
    datum = _to_date(rec.get("Datum"))
    status = _normalize_milestone_status(rec.get("Status"))
    fort = _to_float_or_none(rec.get("Fortschritt"))
    if fort is None:
        fort = 0.0
    fort = max(0.0, min(100.0, float(fort)))
    comment = _safe_str(rec.get("Kommentar")).strip()

    ws.cell(row_idx, 1).value = projekt or None
    ws.cell(row_idx, 2).value = name or None
    ws.cell(row_idx, 3).value = datum
    ws.cell(row_idx, 4).value = status or None
    ws.cell(row_idx, 5).value = float(fort)
    ws.cell(row_idx, 6).value = comment or None

    ws.cell(row_idx, 3).number_format = "DD.MM.YYYY"
    ws.cell(row_idx, 5).number_format = "0"


def _clear_milestone_row(ws, row_idx: int) -> None:
    for c in range(1, 7):
        ws.cell(row_idx, c).value = None


# ------------------------- workbook writing -------------------------

def _find_next_write_row(ws, key_cols=KEY_COLS_FOR_EMPTY_CHECK) -> int:
    for r in range(2, ws.max_row + 2):
        if all(_is_blank(ws.cell(r, c).value) for c in key_cols):
            return r
    return ws.max_row + 1


def _write_taetigkeit_row(ws, row_idx: int, record: Dict[str, Any]) -> None:
    datum = _to_date(record.get("Datum"))
    projekt = _safe_str(record.get("Projekt")).strip()
    zeit_von = _to_time(record.get("Zeit von"))
    zeit_bis = _to_time(record.get("Zeit bis"))
    pause_min = int(record.get("Pause_Min") or 0)
    pause_time = _minutes_to_time(pause_min)
    km = record.get("km")
    try:
        km_val = None if _is_blank(km) else int(float(km))
    except Exception:
        km_val = None
    taet_typ = _safe_str(record.get("Tätigkeit")).strip()
    kodierung = _safe_str(record.get("Kodierung")).strip() or None
    interne = _safe_str(record.get("Interne Projekte")).strip() or None
    info = _safe_str(record.get("Info"))
    abgerechnet = _safe_str(record.get("Abgerechnet")).strip() or None
    eingetragen = _safe_str(record.get("eingetragen")).strip() or None

    ws.cell(row_idx, 1).value = datum
    ws.cell(row_idx, 2).value = projekt or None
    ws.cell(row_idx, 3).value = zeit_von
    ws.cell(row_idx, 4).value = zeit_bis
    ws.cell(row_idx, 5).value = pause_time

    zahl = record.get("Zahl")
    if zeit_von is None and zeit_bis is None and zahl is not None:
        try:
            z_float = float(zahl)
            ws.cell(row_idx, 6).value = z_float / 24.0
            ws.cell(row_idx, 7).value = z_float
        except Exception:
            ws.cell(row_idx, 6).value = ""
            ws.cell(row_idx, 7).value = ""
    else:
        ws.cell(row_idx, 6).value = f'=IF(A{row_idx}="","",D{row_idx}-C{row_idx}-E{row_idx})'
        ws.cell(row_idx, 7).value = f'=IF(A{row_idx}="","",F{row_idx}*24)'

    ws.cell(row_idx, 8).value = km_val
    ws.cell(row_idx, 9).value = taet_typ or None
    ws.cell(row_idx, 10).value = kodierung
    ws.cell(row_idx, 11).value = interne
    ws.cell(row_idx, 12).value = info or None
    ws.cell(row_idx, 13).value = abgerechnet
    ws.cell(row_idx, 14).value = eingetragen

    ws.cell(row_idx, 1).number_format = "DD.MM.YYYY"
    for c in (3, 4, 5, 6):
        ws.cell(row_idx, c).number_format = "hh:mm"
    ws.cell(row_idx, 7).number_format = "0.00"


def _clear_taetigkeit_row(ws, row_idx: int) -> None:
    for c in range(1, 15):
        ws.cell(row_idx, c).value = None


def _ensure_team_sheet(wb: openpyxl.Workbook):
    if TEAM_SHEET in wb.sheetnames:
        return wb[TEAM_SHEET]
    ws = wb.create_sheet(TEAM_SHEET)
    for i, h in enumerate(TEAM_COLS, start=1):
        ws.cell(1, i).value = h
    return ws


def _write_team_row(ws, row_idx: int, record: Dict[str, Any]) -> None:
    ws.cell(row_idx, 1).value = _safe_str(record.get("Mitarbeiter")).strip() or "Unbekannt"
    ws.cell(row_idx, 2).value = _to_date(record.get("Datum"))
    ws.cell(row_idx, 3).value = _safe_str(record.get("Projekt")).strip() or None
    ws.cell(row_idx, 4).value = _to_time(record.get("Zeit von"))
    ws.cell(row_idx, 5).value = _to_time(record.get("Zeit bis"))

    pause_min = int(record.get("Pause_Min") or 0)
    ws.cell(row_idx, 6).value = _minutes_to_time(pause_min)

    zahl = record.get("Zahl")
    if record.get("Zeit von") is None and record.get("Zeit bis") is None and zahl is not None:
        try:
            z_float = float(zahl)
            ws.cell(row_idx, 7).value = z_float / 24.0
            ws.cell(row_idx, 8).value = z_float
        except Exception:
            ws.cell(row_idx, 7).value = ""
            ws.cell(row_idx, 8).value = ""
    else:
        ws.cell(row_idx, 7).value = f'=IF(B{row_idx}="","",E{row_idx}-D{row_idx}-F{row_idx})'
        ws.cell(row_idx, 8).value = f'=IF(B{row_idx}="","",G{row_idx}*24)'

    try:
        km_val = None if _is_blank(record.get("km")) else int(float(record.get("km")))
    except Exception:
        km_val = None

    ws.cell(row_idx, 9).value = km_val
    ws.cell(row_idx, 10).value = _safe_str(record.get("Tätigkeit")).strip() or None
    ws.cell(row_idx, 11).value = _safe_str(record.get("Kodierung")).strip() or None
    ws.cell(row_idx, 12).value = _safe_str(record.get("Interne Projekte")).strip() or None
    ws.cell(row_idx, 13).value = _safe_str(record.get("Info")) or None
    ws.cell(row_idx, 14).value = _safe_str(record.get("Abgerechnet")).strip() or None
    ws.cell(row_idx, 15).value = _safe_str(record.get("eingetragen")).strip() or None

    ws.cell(row_idx, 2).number_format = "DD.MM.YYYY"
    for c in (4, 5, 6, 7):
        ws.cell(row_idx, c).number_format = "hh:mm"
    ws.cell(row_idx, 8).number_format = "0.00"


def _clear_team_row(ws, row_idx: int) -> None:
    for c in range(1, 16):
        ws.cell(row_idx, c).value = None


def _find_hilfstabelle_project_row(ws, project_name: str) -> Optional[int]:
    target = _safe_str(project_name).strip()
    if not target:
        return None
    for r in range(2, ws.max_row + 1):
        p = _safe_str(ws.cell(r, 7).value).strip()
        if p and p == target:
            return r
    return None


def _find_next_hilfstabelle_project_row(ws) -> int:
    for r in range(2, ws.max_row + 2):
        if _is_blank(ws.cell(r, 7).value):
            return r
    return ws.max_row + 1


def _upsert_project_stammdaten(
        wb: openpyxl.Workbook,
        project_data: Dict[str, Any],
        original_project: Optional[str] = None,
        rename_taetigkeiten: bool = False,
) -> Tuple[bool, str]:
    if HILFS_SHEET not in wb.sheetnames:
        return False, f"Arbeitsblatt '{HILFS_SHEET}' nicht gefunden."
    ws = wb[HILFS_SHEET]

    new_project = _safe_str(project_data.get("Projekt")).strip()
    if not new_project:
        return False, "Projektname/-kürzel darf nicht leer sein."

    orig = _safe_str(original_project).strip() if original_project else ""
    row_idx = _find_hilfstabelle_project_row(ws, orig or new_project)

    if row_idx is None:
        existing = _find_hilfstabelle_project_row(ws, new_project)
        if existing is not None:
            row_idx = existing
        else:
            row_idx = _find_next_hilfstabelle_project_row(ws)
    else:
        if orig and orig != new_project:
            conflict = _find_hilfstabelle_project_row(ws, new_project)
            if conflict is not None and conflict != row_idx:
                return False, f"Projekt '{new_project}' existiert bereits in der Hilfstabelle (Zeile {conflict})."

    ws.cell(row_idx, 7).value = new_project
    ws.cell(row_idx, 8).value = _safe_str(project_data.get("Kunde")).strip() or None
    ws.cell(row_idx, 9).value = _safe_str(project_data.get("Straße")).strip() or None
    ws.cell(row_idx, 10).value = _safe_str(project_data.get("Ort")).strip() or None
    ws.cell(row_idx, 11).value = _safe_str(project_data.get("Ansprechpartner")).strip() or None
    ws.cell(row_idx, 12).value = _safe_str(project_data.get("Projektadresse Standard")).strip() or None
    ws.cell(row_idx, 13).value = _safe_str(project_data.get("Projektadresse Alternativ")).strip() or None
    ws.cell(row_idx, 14).value = f'=H{row_idx}&","&" "&I{row_idx}&","&" "&J{row_idx}'

    renamed_count = 0
    if rename_taetigkeiten and orig and orig != new_project and TAETIGKEITEN_SHEET in wb.sheetnames:
        tws = wb[TAETIGKEITEN_SHEET]
        for r in range(2, tws.max_row + 1):
            if _safe_str(tws.cell(r, 2).value).strip() == orig:
                tws.cell(r, 2).value = new_project
                renamed_count += 1

    msg = f"Projektstammdaten gespeichert (Hilfstabelle Zeile {row_idx})."
    if renamed_count:
        msg += f" Tätigkeiten umbenannt: {renamed_count}."
    return True, msg


def _set_relevante_kodierungen(wb: openpyxl.Workbook, selected_aufgaben: List[str]) -> Tuple[bool, str]:
    if KODIERUNG_SHEET not in wb.sheetnames:
        return False, f"Arbeitsblatt '{KODIERUNG_SHEET}' nicht gefunden."
    ws = wb[KODIERUNG_SHEET]
    selected = {_safe_str(x).strip() for x in (selected_aufgaben or []) if _safe_str(x).strip()}
    changed = 0
    total = 0
    for r in range(2, ws.max_row + 1):
        aufgabe = _safe_str(ws.cell(r, 2).value).strip()
        if not aufgabe:
            continue
        total += 1
        new_marker = "x" if aufgabe in selected else None
        old_marker = _safe_str(ws.cell(r, 1).value).strip().lower()
        normalized_old = "x" if old_marker == "x" else ""
        normalized_new = "x" if new_marker == "x" else ""
        if normalized_old != normalized_new:
            ws.cell(r, 1).value = new_marker
            changed += 1
    return True, f"Relevante Kodierungen gespeichert: {len(selected)} ausgewählt ({changed} Änderungen in {total} Kodierungen)."


def _save_workbook(path: Path, mutator) -> Tuple[bool, str]:
    try:
        backup_path = path.with_suffix(path.suffix + f".bak-{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}")
        shutil.copy2(path, backup_path)

        wb = openpyxl.load_workbook(path)
        try:
            if hasattr(wb, "calculation") and wb.calculation is not None:
                wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass
        mutator(wb)
        wb.save(path)
        return True, f"Gespeichert. Backup erstellt: {backup_path.name}"
    except Exception as e:
        return False, f"Fehler beim Speichern: {e}"


# ------------------------- reporting logic -------------------------

def _build_report(df: pd.DataFrame, lookups: Dict[str, Any], year: int, month: int, project: str,
                  include_abgerechnet: bool) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    x = df.copy()
    x["Datum_dt"] = pd.to_datetime(x["Datum"], errors="coerce")
    x = x[x["Datum_dt"].notna()]
    x = x[x["Datum_dt"].dt.year == int(year)]
    x = x[x["Datum_dt"].dt.month == int(month)]
    x = x[x["Projekt"].astype(str) == str(project)]

    x = x[x["Tätigkeit"].astype(str).str.upper() != "I"]

    if not include_abgerechnet:
        x = x[x["Abgerechnet"].fillna("").astype(str).str.strip().str.lower() != "ja"]

    x = x.copy()
    x["Kodierung EB"] = x["Kodierung"].map(lookups.get("kodierung_map_eb", {}))
    x["Datum"] = x["Datum"].apply(_format_date)
    x["Beginn"] = x["Zeit von"].apply(_format_time)
    x["Ende"] = x["Zeit bis"].apply(_format_time)
    x["Pause"] = x["Pause"].apply(_format_time)
    x["Zeit (h)"] = x["Zahl"].apply(lambda v: round(float(v), 2) if v is not None and not pd.isna(v) else None)
    x["Art"] = x["Tätigkeit"]
    x["Leistungsbeschreibung"] = x["Info"].fillna("")

    x["_sort_date"] = pd.to_datetime(x["Datum"], format="%d.%m.%Y", errors="coerce")
    x["_sort_start"] = x["Zeit von"].apply(lambda t: _time_to_minutes(t) if t else -1)
    x = x.sort_values(["_sort_date", "_sort_start", "_excel_row"]).drop(columns=["_sort_date", "_sort_start"])

    return x[
        ["_excel_row", "Datum", "Beginn", "Ende", "Pause", "Zeit (h)", "Art", "Kodierung EB", "Leistungsbeschreibung",
         "Abgerechnet"]].reset_index(drop=True)


def _summaries_from_report(report_df: pd.DataFrame) -> Dict[str, float]:
    if report_df.empty:
        return {"F": 0.0, "R": 0.0, "K": 0.0, "gesamt": 0.0}
    sums = report_df.groupby("Art", dropna=False)["Zeit (h)"].sum(min_count=1).to_dict()
    out = {
        "F": round(float(sums.get("F", 0.0) or 0.0), 2),
        "R": round(float(sums.get("R", 0.0) or 0.0), 2),
        "K": round(float(sums.get("K", 0.0) or 0.0), 2),
    }
    out["gesamt"] = round(out["F"] + out["R"] + out["K"], 2)
    return out


# ------------------------- UI helpers -------------------------

def _project_defaults(df: pd.DataFrame) -> Tuple[int, int]:
    if df.empty or "Datum" not in df.columns:
        today = dt.date.today()
        return today.year, today.month
    dates = [d for d in df["Datum"].tolist() if isinstance(d, dt.date)]
    if not dates:
        today = dt.date.today()
        return today.year, today.month
    latest = max(dates)
    return latest.year, latest.month


def _render_taetigkeit_form(prefix: str, lookups: Dict[str, Any], defaults: Optional[Dict[str, Any]] = None) -> Dict[
    str, Any]:
    defaults = defaults or {}
    projekte = list(dict.fromkeys([*(lookups.get("projekte") or []), _safe_str(defaults.get("Projekt"))]))
    projekte = [p for p in projekte if p]
    typen = list(dict.fromkeys([*(lookups.get("taetigkeit_typen") or []), _safe_str(defaults.get("Tätigkeit"))]))
    ja_nein = list(dict.fromkeys([*(lookups.get("ja_nein") or ["ja", "nein"]), _safe_str(defaults.get("Abgerechnet")),
                                  _safe_str(defaults.get("eingetragen"))]))
    kod_options = list(
        dict.fromkeys(
            [""]
            + (lookups.get("relevante_kodierungen") or [])
            + (lookups.get("kodierung_aufgaben") or [])
            + [_safe_str(defaults.get("Kodierung"))]
        )
    )
    interne_options = list(
        dict.fromkeys([""] + (lookups.get("interne_projekte") or []) + [_safe_str(defaults.get("Interne Projekte"))]))

    d_default = _to_date(defaults.get("Datum")) or dt.date.today()
    zv_default = _to_time(defaults.get("Zeit von")) or dt.time(8, 0)
    zb_default = _to_time(defaults.get("Zeit bis")) or dt.time(9, 0)
    pause_min_default = int(defaults.get("Pause_Min") or _time_to_minutes(defaults.get("Pause")) or 0)
    km_default = 0
    try:
        if not _is_blank(defaults.get("km")):
            km_default = int(float(defaults.get("km")))
    except Exception:
        km_default = 0

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        datum = st.date_input("Datum", value=d_default, key=f"{prefix}_datum")
        projekt = st.selectbox(
            "Projekt",
            options=projekte if projekte else [""],
            index=(projekte.index(_safe_str(defaults.get("Projekt"))) if _safe_str(
                defaults.get("Projekt")) in projekte else 0),
            key=f"{prefix}_projekt",
        )
    with col2:
        zeit_von = st.time_input("Zeit von", value=zv_default, key=f"{prefix}_zv")
        zeit_bis = st.time_input("Zeit bis", value=zb_default, key=f"{prefix}_zb")
    with col3:
        pause_min = st.number_input("Pause (Minuten)", min_value=0, max_value=600, value=pause_min_default, step=5,
                                    key=f"{prefix}_pause")
        km = st.number_input("km", min_value=0, value=km_default, step=1, key=f"{prefix}_km")
    with col4:
        taet_typ = st.selectbox(
            "Tätigkeit (Typ)",
            options=typen if typen else ["F", "R", "I"],
            index=(typen.index(_safe_str(defaults.get("Tätigkeit"))) if _safe_str(
                defaults.get("Tätigkeit")) in typen else 0),
            key=f"{prefix}_typ",
        )
        kodierung = st.selectbox(
            "Kodierung (Aufgabe)",
            options=kod_options,
            index=(kod_options.index(_safe_str(defaults.get("Kodierung"))) if _safe_str(
                defaults.get("Kodierung")) in kod_options else 0),
            key=f"{prefix}_kod",
        )

    col5, col6, col7 = st.columns([2, 1, 1])
    with col5:
        info = st.text_area("Info / Leistungsbeschreibung", value=_safe_str(defaults.get("Info")), height=80,
                            key=f"{prefix}_info")
    with col6:
        interne = st.selectbox(
            "Interne Projekte",
            options=interne_options,
            index=(interne_options.index(_safe_str(defaults.get("Interne Projekte"))) if _safe_str(
                defaults.get("Interne Projekte")) in interne_options else 0),
            key=f"{prefix}_intern",
        )
    with col7:
        abgerechnet = st.selectbox(
            "Abgerechnet",
            options=ja_nein if ja_nein else ["ja", "nein"],
            index=((ja_nein.index(_safe_str(defaults.get("Abgerechnet"))) if _safe_str(
                defaults.get("Abgerechnet")) in ja_nein else (ja_nein.index("nein") if "nein" in ja_nein else 0))),
            key=f"{prefix}_abg",
        )
        eingetragen = st.selectbox(
            "eingetragen",
            options=[""] + (ja_nein if ja_nein else ["ja", "nein"]),
            index=([""] + (ja_nein if ja_nein else ["ja", "nein"])).index(
                _safe_str(defaults.get("eingetragen"))) if _safe_str(defaults.get("eingetragen")) in (
                    [""] + (ja_nein if ja_nein else ["ja", "nein"])) else 0,
            key=f"{prefix}_eing",
        )

    hours = _compute_hours_decimal(zeit_von, zeit_bis, int(pause_min))
    if hours is None:
        st.warning("Zeitangaben ergeben keine gültige Dauer.")
    else:
        st.caption(
            f"Berechnete Dauer: **{hours:.2f} h** ({int(round(hours * 60)) // 60:02d}:{int(round(hours * 60)) % 60:02d})")

    return {
        "Datum": datum,
        "Projekt": projekt,
        "Zeit von": zeit_von,
        "Zeit bis": zeit_bis,
        "Pause_Min": int(pause_min),
        "km": int(km),
        "Tätigkeit": taet_typ,
        "Kodierung": kodierung,
        "Interne Projekte": interne,
        "Info": info,
        "Abgerechnet": abgerechnet,
        "eingetragen": eingetragen,
    }


def _filtered_taetigkeiten(df: pd.DataFrame, year: Optional[int], month: Optional[int], project: str,
                           include_abgerechnet: bool) -> pd.DataFrame:
    x = df.copy()
    if x.empty:
        return x
    if year:
        x = x[x["Datum"].apply(lambda d: isinstance(d, dt.date) and d.year == year)]
    if month:
        x = x[x["Datum"].apply(lambda d: isinstance(d, dt.date) and d.month == month)]
    if project:
        x = x[x["Projekt"].astype(str) == project]
    if not include_abgerechnet:
        x = x[x["Abgerechnet"].fillna("").astype(str).str.strip().str.lower() != "ja"]
    return x.reset_index(drop=True)


def _row_key_from_series(s: pd.Series) -> tuple:
    return (
        _format_date(s.get("Datum")),
        _safe_str(s.get("Projekt")).strip(),
        _format_time(s.get("Zeit von")),
        _format_time(s.get("Zeit bis")),
        int(s.get("Pause_Min") or _time_to_minutes(s.get("Pause")) or 0),
        int(float(s.get("km") or 0) or 0),
        _safe_str(s.get("Tätigkeit")).strip(),
        _safe_str(s.get("Kodierung")).strip(),
        _safe_str(s.get("Interne Projekte")).strip(),
        _safe_str(s.get("Info")).strip(),
    )


def _existing_keys(df: pd.DataFrame) -> set:
    if df is None or df.empty:
        return set()
    keys = set()
    for _, r in df.iterrows():
        keys.add(_row_key_from_series(r))
    return keys


@st.cache_data(show_spinner=False, max_entries=100)
def _parse_and_store_uploaded_report(file_name: str, file_bytes: bytes, allowed_types: Tuple[str, ...]) -> Tuple[
    Dict[str, Any], pd.DataFrame]:
    h = hashlib.md5(file_bytes).hexdigest()[:8]
    script_dir = Path(__file__).resolve().parent
    reports_dir = script_dir / "imports_reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    stem = Path(file_name).stem
    target = reports_dir / f"{stem}_{h}.xlsx"
    if not target.exists():
        target.write_bytes(file_bytes)

    meta, lines = _read_einsatzbericht_xlsx(target, list(allowed_types))
    meta["original_filename"] = file_name
    return meta, lines


def _guess_project(meta: Dict[str, Any], filename: str, lookups: Dict[str, Any], target_projects: List[str]) -> str:
    det = _safe_str(meta.get("project")).strip()
    fname_lower = filename.lower()

    if det in target_projects:
        return det

    for p in sorted(target_projects, key=len, reverse=True):
        if p and p.lower() in fname_lower:
            return p

    infos = lookups.get("projekt_infos", {})
    if det:
        det_lower = det.lower()
        for p, p_data in infos.items():
            kunde = _safe_str(p_data.get("Kunde")).strip().lower()
            if kunde and (kunde in det_lower or det_lower in kunde):
                return p

    for p, p_data in infos.items():
        kunde = _safe_str(p_data.get("Kunde")).strip().lower()
        if kunde and kunde in fname_lower:
            return p

    return ""


def _build_reverse_kod_map_eb_to_aufgabe(lookups: Dict[str, Any]) -> Dict[str, str]:
    fwd = lookups.get("kodierung_map_eb", {}) or {}
    rev: Dict[str, str] = {}
    collisions = set()
    for aufgabe, eb in fwd.items():
        eb_s = _safe_str(eb).strip()
        aufg_s = _safe_str(aufgabe).strip()
        if not eb_s or not aufg_s:
            continue
        if eb_s in rev and rev[eb_s] != aufg_s:
            collisions.add(eb_s)
        else:
            rev[eb_s] = aufg_s
    for c in collisions:
        rev.pop(c, None)
    return rev


GER_MONTHS = {
    "januar": 1, "februar": 2, "märz": 3, "maerz": 3, "april": 4, "mai": 5, "juni": 6,
    "juli": 7, "august": 8, "september": 9, "oktober": 10, "november": 11, "dezember": 12,
}


def _parse_month_year_from_filename(p: Path) -> Tuple[Optional[int], Optional[int]]:
    s = p.stem.lower()
    y = None
    m = None

    my = re.search(r"\b(20\d{2})\b", s)
    if my:
        y = int(my.group(1))

    mm = re.search(r"\b(januar|februar|märz|maerz|april|mai|juni|juli|august|september|oktober|november|dezember)\b", s)
    if mm:
        m = GER_MONTHS.get(mm.group(1))
    return y, m


def _read_project_from_sheet(ws) -> str:
    c12 = _safe_str(ws["C12"].value).strip()
    if c12:
        return c12
    for r in range(1, 50):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "firma:":
                right = _safe_str(ws.cell(r, c + 1).value).strip()
                if right:
                    return right
    return ""


def _read_name_from_sheet(ws) -> str:
    # 1. Oftmals in C7 hinterlegt
    c7 = _safe_str(ws["C7"].value).strip()

    # 2. Sicherheitshalber dynamisch scannen
    for r in range(1, 20):
        for c in range(1, 10):
            v = _safe_str(ws.cell(r, c).value).strip().lower()
            if "berater" in v or "name" in v or "mitarbeiter" in v:
                for offset in range(1, 4):
                    val = _safe_str(ws.cell(r, c + offset).value).strip()
                    if val and len(val) > 2:
                        return val
    if c7 and len(c7) > 2:
        return c7
    return ""


def _find_header(ws, max_rows: int = 80, max_cols: int = 30) -> Tuple[Optional[int], Dict[str, int]]:
    required = {"datum", "beginn", "ende", "art"}
    for r in range(1, max_rows + 1):
        mapping: Dict[str, int] = {}
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                mapping[v.strip().lower()] = c
        if required.issubset(mapping.keys()):
            return r, mapping
    return None, {}


def _read_einsatzbericht_xlsx(path: Path, allowed_types: List[str]) -> Tuple[Dict[str, Any], pd.DataFrame]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for sheet_name in ["Einsatzbericht", "Tabelle1"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    header_row, col = _find_header(ws)
    if header_row is None:
        meta = {"project": "", "year": None, "month": None, "source_path": str(path), "name": ""}
        return meta, pd.DataFrame(
            columns=["Datum", "Beginn", "Ende", "Pause_Min", "Zeit_h", "Art", "Kodierung_EB", "Leistungsbeschreibung"])

    project = _read_project_from_sheet(ws)
    mitarbeiter_name = _read_name_from_sheet(ws)

    year = ws["K2"].value
    month = ws["K3"].value
    if _is_blank(year) or _is_blank(month):
        y2, m2 = _parse_month_year_from_filename(path)
        year = year if not _is_blank(year) else y2
        month = month if not _is_blank(month) else m2

    def c(name: str) -> Optional[int]:
        return col.get(name.lower())

    c_datum = c("datum")
    c_beginn = c("beginn")
    c_ende = c("ende")
    c_pause = c("pause (h)") or c("pause")
    c_zeit = c("zeit  (h)") or c("zeit (h)") or c("zeit")
    c_art = c("art")
    c_text = c("leistungsbeschreibung") or c("beschreibung")
    c_kod = c("kodierung") or c("kodierung eb")

    rows = []
    empty_streak = 0
    start = header_row + 1
    last_seen_date = None

    valid_arts = {str(t).strip().lower() for t in allowed_types if str(t).strip()}

    for r in range(start, start + 400):
        a = ws.cell(r, c_datum).value if c_datum else None
        b = ws.cell(r, c_beginn).value if c_beginn else None
        e = ws.cell(r, c_ende).value if c_ende else None
        t = ws.cell(r, c_text).value if c_text else None
        z_val = ws.cell(r, c_zeit).value if c_zeit else None
        art_val = ws.cell(r, c_art).value if c_art else None

        if _is_blank(a) and _is_blank(b) and _is_blank(e) and _is_blank(t) and _is_blank(z_val) and _is_blank(art_val):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        art = _safe_str(art_val).strip()

        if not art or art.lower() not in valid_arts:
            continue

        datum = _to_date(a)
        if datum is not None:
            last_seen_date = datum
        elif last_seen_date is not None:
            datum = last_seen_date

        beginn = _to_time(b)
        ende = _to_time(e)
        text = _safe_str(t).strip()

        zeit_h = None
        if not _is_blank(z_val):
            if isinstance(z_val, (int, float)):
                zeit_h = float(z_val)
            elif isinstance(z_val, (dt.time, dt.datetime)):
                zeit_h = z_val.hour + z_val.minute / 60.0
            else:
                try:
                    zeit_h = float(str(z_val).strip().replace(',', '.'))
                except Exception:
                    zeit_h = None

        kod_eb = _safe_str(ws.cell(r, c_kod).value).strip() if c_kod else ""

        has_manual_content = False
        if zeit_h is not None:
            try:
                has_manual_content = abs(float(zeit_h)) > 1e-9
            except Exception:
                has_manual_content = True

        if beginn is None and ende is None and not text and not has_manual_content:
            continue

        pause_min = 0
        if c_pause:
            p_val = ws.cell(r, c_pause).value
            if not _is_blank(p_val):
                if isinstance(p_val, (int, float)):
                    pause_min = int(round(float(p_val) * 60))
                elif isinstance(p_val, (dt.time, dt.datetime)):
                    pause_min = p_val.hour * 60 + p_val.minute
                elif isinstance(p_val, str):
                    s_val = p_val.strip().replace(',', '.')
                    try:
                        pause_min = int(round(float(s_val) * 60))
                    except ValueError:
                        pause_min = _time_to_minutes(_to_time(p_val))

        if beginn is not None and ende is not None and zeit_h is not None:
            total_duration_hours = _compute_hours_decimal(beginn, ende, 0)
            if total_duration_hours is not None:
                inferred_pause_hours = total_duration_hours - zeit_h
                inferred_pause_min = int(round(inferred_pause_hours * 60))
                if -5 <= inferred_pause_min < 720:
                    pause_min = max(0, inferred_pause_min)

        if pause_min >= 720:
            pause_min = 0

        if zeit_h is None and beginn is not None and ende is not None:
            zeit_h = _compute_hours_decimal(beginn, ende, pause_min)

        rows.append({
            "Datum": datum,
            "Beginn": beginn,
            "Ende": ende,
            "Pause_Min": pause_min,
            "Zeit_h": zeit_h,
            "Art": art,
            "Kodierung_EB": kod_eb,
            "Leistungsbeschreibung": text,
            "_source_row": r,
        })

    df_lines = pd.DataFrame(rows)
    if (year is None or month is None) and not df_lines.empty:
        first_date = df_lines["Datum"].dropna().iloc[0] if df_lines["Datum"].notna().any() else None
        if isinstance(first_date, dt.date):
            year = year or first_date.year
            month = month or first_date.month

    meta = {
        "project": project,
        "year": year,
        "month": month,
        "name": mitarbeiter_name,
        "source_path": str(path),
    }
    return meta, df_lines


def _key_for_import(rec: Dict[str, Any]) -> tuple:
    d = _format_date(rec.get("Datum"))
    projekt = _safe_str(rec.get("Projekt")).strip()
    zv = _format_time(rec.get("Zeit von"))
    zb = _format_time(rec.get("Zeit bis"))
    info = _safe_str(rec.get("Info")).strip()
    typ = _safe_str(rec.get("Tätigkeit")).strip()
    kod = _safe_str(rec.get("Kodierung")).strip()
    pause = int(rec.get("Pause_Min") or 0)

    if not zv and not zb:
        zeit_h = rec.get("Zeit_h") or rec.get("Zeit (h)") or rec.get("Zahl")
        zeit_h_norm = ""
        if zeit_h is not None and not _is_blank(zeit_h):
            try:
                zeit_h_norm = f"{float(zeit_h):.4f}"
            except Exception:
                zeit_h_norm = _safe_str(zeit_h).strip()
        return (d, projekt, "<NO_TIME>", typ, kod, zeit_h_norm, info)

    return (d, projekt, zv, zb, pause, typ, kod, info)


def _existing_keys_for_master(df_master: pd.DataFrame, team: bool = False) -> set:
    keys = set()
    if df_master is None or df_master.empty:
        return keys
    for _, r in df_master.iterrows():
        rec = {
            "Datum": r.get("Datum"),
            "Projekt": r.get("Projekt"),
            "Zeit von": r.get("Zeit von"),
            "Zeit bis": r.get("Zeit bis"),
            "Pause_Min": int(r.get("Pause_Min") or _time_to_minutes(r.get("Pause")) or 0),
            "km": r.get("km") or 0,
            "Tätigkeit": r.get("Tätigkeit"),
            "Kodierung": r.get("Kodierung") or "",
            "Info": r.get("Info") or "",
        }
        if team:
            key_tuple = (_safe_str(r.get("Mitarbeiter")).strip(),) + _key_for_import(rec)
            keys.add(key_tuple)
        else:
            keys.add(_key_for_import(rec))
    return keys


# ------------------------- Streamlit UI -------------------------

def _to_float_or_none(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def _viz_base_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=["Datum_dt", "Year", "Month", "YM", "YM_dt", "Projekt", "Tätigkeit", "Hours", "Abgerechnet", "km",
                     "Kodierung", "Info", "Mitarbeiter"])

    x = df.copy()
    x["Datum_dt"] = pd.to_datetime(x["Datum"], errors="coerce")
    x = x[x["Datum_dt"].notna()].copy()

    x["Year"] = x["Datum_dt"].dt.year.astype(int)
    x["Month"] = x["Datum_dt"].dt.month.astype(int)

    per = x["Datum_dt"].dt.to_period("M")
    x["YM"] = per.astype(str)
    x["YM_dt"] = per.dt.to_timestamp()

    x["Hours"] = x["Zahl"].apply(_to_float_or_none)
    x["Hours"] = pd.to_numeric(x["Hours"], errors="coerce").fillna(0.0)

    x["Projekt"] = x["Projekt"].astype(str).fillna("").str.strip()
    x["Tätigkeit"] = x["Tätigkeit"].astype(str).fillna("").str.strip().str.upper()

    x["Abgerechnet"] = x["Abgerechnet"].fillna("").astype(str).str.strip().str.lower()
    x["Abgerechnet"] = x["Abgerechnet"].map(
        lambda s: "ja" if s in {"ja", "yes", "y", "true", "1"} else (
            "nein" if s in {"nein", "no", "n", "false", "0"} else s)
    )
    x["km"] = pd.to_numeric(x.get("km"), errors="coerce").fillna(0).astype(int)

    if "Kodierung" not in x.columns:
        x["Kodierung"] = ""
    if "Info" not in x.columns:
        x["Info"] = ""
    if "Mitarbeiter" not in x.columns:
        x["Mitarbeiter"] = "Ich (Eigene)"
    x["Mitarbeiter"] = x["Mitarbeiter"].fillna("Unbekannt")

    return x


def _vega_spec_for_chart(
        kind: str,
        x_field: str,
        y_field: str,
        color_field: Optional[str] = None,
        x_type: str = "ordinal",
        y_type: str = "quantitative",
        stacked: bool = False,
        donut: bool = False,
) -> Dict[str, Any]:
    kind = (kind or "").lower().strip()

    if kind in {"pie", "donut"}:
        inner = 60 if (kind == "donut" or donut) else 0
        spec = {
            "mark": {"type": "arc", "innerRadius": inner},
            "encoding": {
                "theta": {"field": y_field, "type": "quantitative"},
                "color": {"field": x_field, "type": "nominal"},
                "tooltip": [
                    {"field": x_field, "type": "nominal"},
                    {"field": y_field, "type": "quantitative", "format": ".2f"},
                ],
            },
        }
        return spec

    enc: Dict[str, Any] = {
        "x": {"field": x_field, "type": x_type},
        "y": {"field": y_field, "type": y_type},
        "tooltip": [
            {"field": x_field, "type": x_type if x_type != "temporal" else "temporal"},
            {"field": y_field, "type": "quantitative", "format": ".2f"},
        ],
    }

    if color_field:
        enc["color"] = {"field": color_field, "type": "nominal"}
        if stacked:
            enc["y"]["stack"] = "zero"

        enc["tooltip"] = [
            {"field": x_field, "type": x_type if x_type != "temporal" else "temporal"},
            {"field": color_field, "type": "nominal"},
            {"field": y_field, "type": "quantitative", "format": ".2f"},
        ]

    mark_type = "bar"
    if kind in {"bar", "stacked bar"}:
        mark_type = "bar"
    elif kind in {"line"}:
        mark_type = "line"
    elif kind in {"area", "stacked area"}:
        mark_type = "area"

    mark: Dict[str, Any] = {"type": mark_type}
    if mark_type == "line":
        mark["point"] = True

    spec = {"mark": mark, "encoding": enc}
    return spec


def _render_chart_block(
        *,
        title: str,
        df: pd.DataFrame,
        key_prefix: str,
        default_kind: str,
        allowed_kinds: List[str],
        x_field: str,
        y_field: str,
        color_field: Optional[str] = None,
        x_type: str = "ordinal",
        stacked_default: bool = False,
        pie_ok: bool = True,
) -> None:
    st.markdown(f"### {title}")

    if df is None or df.empty:
        st.caption("Keine Daten.")
        return

    kind_key = f"{key_prefix}_kind"
    if kind_key not in st.session_state:
        st.session_state[kind_key] = default_kind

    c1, c2, c3 = st.columns([2, 1, 2])
    with c1:
        kind = st.selectbox(
            "Diagrammtyp",
            options=allowed_kinds,
            index=allowed_kinds.index(st.session_state[kind_key]) if st.session_state[kind_key] in allowed_kinds else 0,
            key=kind_key,
        )
    with c2:
        stacked = False
        if color_field and kind.lower() in {"stacked bar", "stacked area"}:
            stacked = True
        elif color_field and kind.lower() in {"bar", "area"}:
            stacked = st.checkbox("stacked", value=stacked_default, key=f"{key_prefix}_stacked")
    with c3:
        show_custom = st.checkbox("Custom Vega-Lite JSON", value=False, key=f"{key_prefix}_custom_toggle")

    is_pie = kind.lower() in {"pie", "donut"}
    if is_pie and not pie_ok:
        st.info("Pie/Donut macht für diese Ansicht keinen Sinn. Wähle Bar/Line/Area.")
        return

    default_spec = _vega_spec_for_chart(
        kind=kind,
        x_field=x_field if not is_pie else (color_field or x_field),
        y_field=y_field,
        color_field=(color_field if not is_pie else None),
        x_type=x_type if not is_pie else "nominal",
        stacked=stacked,
        donut=(kind.lower() == "donut"),
    )

    spec_key = f"{key_prefix}_custom_spec"
    if spec_key not in st.session_state:
        st.session_state[spec_key] = json.dumps(default_spec, ensure_ascii=False, indent=2)

    if show_custom:
        raw = st.text_area(
            "Vega-Lite Spec (JSON)",
            value=st.session_state[spec_key],
            height=260,
            key=f"{key_prefix}_spec_editor",
            help="Hier kannst du die Vega-Lite JSON Spezifikation anpassen.",
        )
        try:
            spec = json.loads(raw)
            st.session_state[spec_key] = raw
        except Exception as e:
            st.error(f"Ungültiges JSON: {e}")
            spec = default_spec
    else:
        st.session_state[spec_key] = json.dumps(default_spec, ensure_ascii=False, indent=2)
        spec = default_spec

    st.vega_lite_chart(
        data=df,
        spec=spec,
        use_container_width=True,
    )


def _render_visualisierung_tab(df: pd.DataFrame, team_df: pd.DataFrame, lookups: Dict[str, Any], xlsx_path: Path,
                               milestones_df: pd.DataFrame) -> None:
    st.subheader("Visualisierung / Controlling (Inkl. Team-Daten)")

    # Combine my data and team data
    my_df = df.copy()
    my_df["Mitarbeiter"] = "Ich (Eigene)"
    if not team_df.empty:
        t_df = team_df.copy()
        t_df["Mitarbeiter"] = t_df["Mitarbeiter"].fillna("Unbekannt")
        combined_df = pd.concat([my_df, t_df], ignore_index=True)
    else:
        combined_df = my_df

    base = _viz_base_df(combined_df)
    if base.empty:
        st.info("Keine Daten vorhanden.")
        return

    all_projects = sorted([p for p in base["Projekt"].unique().tolist() if p and p != "nan"])
    all_years = sorted(base["Year"].unique().tolist())
    all_mitarbeiter = sorted([m for m in base["Mitarbeiter"].unique().tolist() if m and m != "nan"])

    min_y = min(all_years) if all_years else dt.date.today().year
    max_y = max(all_years) if all_years else dt.date.today().year

    c1, c2, c3, c4 = st.columns([2, 1, 1, 2])
    with c1:
        sel_projects = st.multiselect(
            "Projekte",
            options=all_projects,
            default=all_projects[:1] if all_projects else [],
            help="Mehrere Projekte auswählen, um Vergleich/Portfolio zu sehen."
        )
    with c2:
        y_from = st.number_input("Jahr von", min_value=2000, max_value=2100, value=int(min_y), step=1, key="viz_y_from")
    with c3:
        y_to = st.number_input("Jahr bis", min_value=2000, max_value=2100, value=int(max_y), step=1, key="viz_y_to")
    with c4:
        sel_mitarbeiter = st.multiselect(
            "Mitarbeiter",
            options=all_mitarbeiter,
            default=all_mitarbeiter,
            help="Welche Personen sollen in die Auswertung/das Budget einfließen?"
        )
        include_abg = st.checkbox("abgerechnete einschließen", value=True, key="viz_include_abg")
        include_internal = st.checkbox("interne Tätigkeiten (I) einschließen", value=True, key="viz_include_internal")

    x = base.copy()
    x = x[(x["Year"] >= int(y_from)) & (x["Year"] <= int(y_to))]
    if sel_projects:
        x = x[x["Projekt"].isin(sel_projects)]
    if sel_mitarbeiter:
        x = x[x["Mitarbeiter"].isin(sel_mitarbeiter)]
    if not include_abg:
        x = x[x["Abgerechnet"] != "ja"]
    if not include_internal:
        x = x[x["Tätigkeit"] != "I"]

    if x.empty:
        st.warning("Keine Daten für diese Auswahl.")
        return

    st.markdown("---")
    st.subheader("Zeitschiene / Meilensteine")

    today = dt.date.today()

    if not isinstance(milestones_df, pd.DataFrame) or milestones_df.empty:
        ms_all = pd.DataFrame(columns=MILESTONE_COLS + ["_excel_row"])
    else:
        ms_all = milestones_df.copy()

    if len(sel_projects) != 1:
        st.info("Für die Zeitschiene wähle bitte **genau ein Projekt** aus.")
    else:
        proj = sel_projects[0]
        ms_proj = ms_all[ms_all["Projekt"].astype(str).str.strip() == str(proj)].copy()
        if ms_proj.empty:
            ms_proj = pd.DataFrame(columns=MILESTONE_COLS + ["_excel_row"])

        ms_kpi = ms_proj.copy()
        ms_kpi["Datum_dt"] = pd.to_datetime(ms_kpi["Datum"], errors="coerce")
        ms_kpi["Status"] = ms_kpi["Status"].apply(_normalize_milestone_status)
        ms_kpi["Fortschritt"] = ms_kpi["Fortschritt"].apply(
            lambda v: max(0.0, min(100.0, float(_to_float_or_none(v) or 0.0))))

        overdue = 0
        done = 0
        next_ms_name = "-"
        next_ms_date = None
        if not ms_kpi.empty and ms_kpi["Datum_dt"].notna().any():
            done = int((ms_kpi["Status"] == "erledigt").sum())
            overdue = int(((ms_kpi["Datum_dt"].dt.date < today) & (ms_kpi["Status"] != "erledigt")).sum())
            upcoming = ms_kpi[ms_kpi["Datum_dt"].dt.date >= today].sort_values("Datum_dt")
            if not upcoming.empty:
                next_ms_name = _safe_str(upcoming.iloc[0].get("Meilenstein"))
                next_ms_date = upcoming.iloc[0]["Datum_dt"].date()

        overall = 0.0
        if not ms_kpi.empty:
            if ms_kpi["Fortschritt"].notna().any():
                overall = float(ms_kpi["Fortschritt"].mean())
            else:
                overall = 100.0 * (done / max(1, len(ms_kpi)))

        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Meilensteine", f"{len(ms_kpi)}")
        a2.metric("Erledigt", f"{done}")
        a3.metric("Überfällig", f"{overdue}")
        a4.metric("Gesamtfortschritt", f"{overall:.1f}%")

        if next_ms_date:
            st.caption(f"Nächster Meilenstein: **{next_ms_name}** am **{next_ms_date.strftime('%d.%m.%Y')}**")
        else:
            st.caption("Kein kommender Meilenstein erkannt (oder keine Datumsspalte gepflegt).")

        if overall > 0:
            st.progress(min(max(overall / 100.0, 0.0), 1.0))

        st.markdown("### Meilensteine pflegen")
        st.caption("Die Tabelle wird im Excel-Sheet **'Meilensteine'** gespeichert.")

        editor_cols = ["_excel_row", "Datum", "Meilenstein", "Status", "Fortschritt", "Kommentar", "Löschen"]
        ed = ms_proj.copy()

        for c in ["Datum", "Meilenstein", "Status", "Fortschritt", "Kommentar", "_excel_row"]:
            if c not in ed.columns:
                ed[c] = None

        ed["Status"] = ed["Status"].apply(_normalize_milestone_status)
        ed["Fortschritt"] = ed["Fortschritt"].apply(lambda v: max(0.0, min(100.0, float(_to_float_or_none(v) or 0.0))))
        ed["Löschen"] = False

        data_editor_fn = getattr(st, "data_editor", None) or getattr(st, "experimental_data_editor")

        edited_ms = data_editor_fn(
            ed[editor_cols],
            key=f"milestones_editor_{proj}",
            use_container_width=True,
            height=320,
            num_rows="dynamic",
            hide_index=True,
            disabled=["_excel_row"],
            column_config={
                "_excel_row": st.column_config.NumberColumn("Excel-Zeile", help="Technische ID (nicht ändern)"),
                "Datum": st.column_config.DateColumn("Datum", format="DD.MM.YYYY"),
                "Meilenstein": st.column_config.TextColumn("Meilenstein", width="large"),
                "Status": st.column_config.SelectboxColumn("Status", options=MILESTONE_STATUSES),
                "Fortschritt": st.column_config.NumberColumn("Fortschritt (%)", min_value=0.0, max_value=100.0,
                                                             step=5.0),
                "Kommentar": st.column_config.TextColumn("Kommentar", width="large"),
                "Löschen": st.column_config.CheckboxColumn("Löschen"),
            },
        )

        if st.button("Meilensteine speichern", key=f"save_milestones_{proj}"):
            orig_by_row: Dict[int, Dict[str, Any]] = {}
            original_excel_rows = set()
            for _, r in ed.iterrows():
                exr = r.get("_excel_row")
                if exr is None or (isinstance(exr, float) and pd.isna(exr)):
                    continue
                try:
                    row_idx = int(exr)
                    orig_by_row[row_idx] = {
                        "Projekt": proj,
                        "Datum": _to_date(r.get("Datum")),
                        "Meilenstein": _safe_str(r.get("Meilenstein")),
                        "Status": _normalize_milestone_status(r.get("Status")),
                        "Fortschritt": float(_to_float_or_none(r.get("Fortschritt")) or 0.0),
                        "Kommentar": _safe_str(r.get("Kommentar")),
                    }
                    original_excel_rows.add(row_idx)
                except Exception:
                    pass

            updates: List[Tuple[int, Dict[str, Any]]] = []
            inserts: List[Dict[str, Any]] = []
            deletes: List[int] = []
            kept_excel_rows = set()

            def _is_blank_ms_row(rr: pd.Series) -> bool:
                return _is_blank(rr.get("Datum")) and _is_blank(rr.get("Meilenstein")) and _is_blank(
                    rr.get("Kommentar"))

            for _, r in edited_ms.iterrows():
                exr = r.get("_excel_row")
                mark_delete = bool(r.get("Löschen", False))
                is_new = (exr is None) or (isinstance(exr, float) and pd.isna(exr))

                if is_new:
                    if _is_blank_ms_row(r):
                        continue
                    rec = {
                        "Projekt": proj,
                        "Datum": _to_date(r.get("Datum")),
                        "Meilenstein": _safe_str(r.get("Meilenstein")).strip(),
                        "Status": _normalize_milestone_status(r.get("Status")),
                        "Fortschritt": float(_to_float_or_none(r.get("Fortschritt")) or 0.0),
                        "Kommentar": _safe_str(r.get("Kommentar")).strip(),
                    }
                    if rec["Meilenstein"]:
                        inserts.append(rec)
                    continue

                row_excel = int(exr)
                kept_excel_rows.add(row_excel)

                if mark_delete:
                    deletes.append(row_excel)
                    continue

                new_rec = {
                    "Projekt": proj,
                    "Datum": _to_date(r.get("Datum")),
                    "Meilenstein": _safe_str(r.get("Meilenstein")).strip(),
                    "Status": _normalize_milestone_status(r.get("Status")),
                    "Fortschritt": float(_to_float_or_none(r.get("Fortschritt")) or 0.0),
                    "Kommentar": _safe_str(r.get("Kommentar")).strip(),
                }
                old_rec = orig_by_row.get(row_excel)
                if old_rec is None or new_rec != old_rec:
                    updates.append((row_excel, new_rec))

            for r_id in original_excel_rows:
                if r_id not in kept_excel_rows and r_id not in deletes:
                    deletes.append(r_id)

            def _mutator_ms(wb):
                ws = _ensure_milestones_sheet(wb)
                for r in deletes:
                    _clear_milestone_row(ws, r)
                for r, rec in updates:
                    _write_milestone_row(ws, r, rec)
                row_idx = _find_next_milestone_row(ws)
                for rec in inserts:
                    _write_milestone_row(ws, row_idx, rec)
                    row_idx += 1

            ok, msg = _save_workbook(Path(xlsx_path), _mutator_ms)
            if ok:
                st.success(
                    f"Meilensteine gespeichert. Updates: {len(updates)}, Neu: {len(inserts)}, Gelöscht: {len(deletes)}. {msg}")
                st.cache_data.clear()
                st.rerun()
            else:
                st.error(msg)

        plot = ms_kpi.copy()
        if plot.empty:
            st.info("Noch keine Meilensteine für dieses Projekt gepflegt.")
        else:
            plot = plot[plot["Datum_dt"].notna()].copy()
            if plot.empty:
                st.info("Meilensteine vorhanden, aber ohne gültiges Datum.")
            else:
                plot["Meilenstein"] = plot["Meilenstein"].fillna("").astype(str)
                plot["Status"] = plot["Status"].apply(_normalize_milestone_status)
                plot["Fortschritt"] = plot["Fortschritt"].apply(
                    lambda v: max(0.0, min(100.0, float(_to_float_or_none(v) or 0.0)))
                )

                st.markdown("### Zeitschiene")
                today_str = dt.date.today().strftime("%Y-%m-%d")

                timeline_spec = {
                    "layer": [
                        {
                            "mark": {"type": "rule", "strokeDash": [6, 6]},
                            "encoding": {
                                "x": {"datum": today_str, "type": "temporal"},
                                "tooltip": [{"datum": today_str, "type": "temporal", "title": "Heute"}],
                            },
                        },
                        {
                            "mark": {"type": "point", "filled": True, "size": 80},
                            "encoding": {
                                "x": {"field": "Datum_dt", "type": "temporal"},
                                "y": {"field": "Meilenstein", "type": "nominal", "sort": None},
                                "color": {"field": "Status", "type": "nominal"},
                                "tooltip": [
                                    {"field": "Meilenstein", "type": "nominal"},
                                    {"field": "Datum_dt", "type": "temporal", "title": "Datum"},
                                    {"field": "Status", "type": "nominal"},
                                    {"field": "Fortschritt", "type": "quantitative", "format": ".0f"},
                                    {"field": "Kommentar", "type": "nominal"},
                                ],
                            },
                        },
                        {
                            "mark": {"type": "text", "align": "left", "dx": 7, "dy": -7},
                            "encoding": {
                                "x": {"field": "Datum_dt", "type": "temporal"},
                                "y": {"field": "Meilenstein", "type": "nominal", "sort": None},
                                "text": {"field": "Fortschritt", "type": "quantitative", "format": ".0f"},
                            },
                        },
                    ],
                }

                st.vega_lite_chart(
                    data=plot,
                    spec=timeline_spec,
                    use_container_width=True,
                )

    budgets = st.session_state.setdefault("viz_budget_eur_by_project", {})
    rates = st.session_state.setdefault("viz_rates_by_project", {})

    st.markdown("### Budget / Stundensätze (optional)")
    st.caption("Alle gewählten Mitarbeiter (siehe Filter oben) fließen kumuliert in diesen Budgetverbrauch mit ein.")

    def _get_rate_map(p: str) -> Dict[str, float]:
        rm = rates.get(p) or {}
        return {
            "default": float(rm.get("default", 0.0) or 0.0),
            "F": float(rm.get("F", 0.0) or 0.0),
            "R": float(rm.get("R", 0.0) or 0.0),
            "S": float(rm.get("S", 0.0) or 0.0),
            "K": float(rm.get("K", 0.0) or 0.0),
            "I": float(rm.get("I", 0.0) or 0.0),
        }

    if len(sel_projects) == 1:
        p = sel_projects[0]
        rm = _get_rate_map(p)

        cc1, cc2, cc3, cc4, cc5, cc6 = st.columns([1, 1, 1, 1, 1, 1])
        with cc1:
            b = st.number_input("Budget (€)", min_value=0.0, value=float(budgets.get(p, 0.0) or 0.0), step=100.0,
                                key=f"viz_budget_{p}")
        with cc2:
            r_def = st.number_input("Rate Default (€/h)", min_value=0.0, value=float(rm["default"]), step=1.0,
                                    key=f"viz_rate_def_{p}")
        with cc3:
            r_f = st.number_input("Rate F (€/h)", min_value=0.0, value=float(rm["F"]), step=1.0, key=f"viz_rate_F_{p}")
        with cc4:
            r_r = st.number_input("Rate R (€/h)", min_value=0.0, value=float(rm["R"]), step=1.0, key=f"viz_rate_R_{p}")
        with cc5:
            r_s = st.number_input("Rate S (€/h)", min_value=0.0, value=float(rm["S"]), step=1.0, key=f"viz_rate_S_{p}")
        with cc6:
            r_k = st.number_input("Rate K (€/h)", min_value=0.0, value=float(rm["K"]), step=1.0, key=f"viz_rate_K_{p}")

        budgets[p] = float(b or 0.0)
        rates[p] = {"default": float(r_def), "F": float(r_f), "R": float(r_r), "S": float(r_s), "K": float(r_k),
                    "I": float(r_def)}
    else:
        st.info("Für Budgetverbrauch wähle genau **ein** Projekt aus (dann kannst du Budget & Stundensätze pflegen).")

    hours_by_type = x.groupby("Tätigkeit")["Hours"].sum().to_dict()
    total_hours = float(sum(hours_by_type.values()) or 0.0)
    total_km = int(x["km"].sum())

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Stunden gesamt", f"{total_hours:.2f} h")
    m2.metric("F", f"{float(hours_by_type.get('F', 0.0)):.2f} h")
    m3.metric("R", f"{float(hours_by_type.get('R', 0.0)):.2f} h")
    m4.metric("S", f"{float(hours_by_type.get('S', 0.0)):.2f} h")
    m5.metric("K", f"{float(hours_by_type.get('K', 0.0)):.2f} h")
    m6.metric("km", f"{total_km}")

    if len(sel_projects) == 1:
        p = sel_projects[0]
        rm = _get_rate_map(p)

        def _rate_for(t: str) -> float:
            t = (t or "").upper()
            rt = float(rm.get(t, 0.0) or 0.0)
            return rt if rt > 0 else float(rm.get("default", 0.0) or 0.0)

        cost = 0.0
        for t, h in hours_by_type.items():
            cost += float(h or 0.0) * _rate_for(t)

        budget = float(budgets.get(p, 0.0) or 0.0)
        st.markdown("### Budgetstatus (Gesamtteam)")
        cst1, cst2, cst3 = st.columns(3)
        fmt = lambda v: f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        cst1.metric("Kosten (Schätzung)", fmt(cost))
        cst2.metric("Budget", fmt(budget))
        remaining = budget - cost
        cst3.metric("Rest", fmt(remaining))

        if budget > 0:
            used_pct = min(max(cost / budget, 0.0), 1.0)
            st.progress(used_pct)
            st.caption(f"Budgetverbrauch: {used_pct * 100:.1f}%")

    monthly_long = (
        x.groupby(["YM_dt", "YM", "Tätigkeit"], as_index=False)["Hours"]
        .sum()
        .sort_values(["YM_dt", "Tätigkeit"])
    )
    proj_df = x.groupby("Projekt", as_index=False)["Hours"].sum().sort_values("Hours", ascending=False)
    type_df = x.groupby("Tätigkeit", as_index=False)["Hours"].sum().sort_values("Hours", ascending=False)
    mitarbeiter_df = x.groupby("Mitarbeiter", as_index=False)["Hours"].sum().sort_values("Hours", ascending=False)

    _render_chart_block(
        title="Stundenverlauf (Monate)",
        df=monthly_long,
        key_prefix="viz_monthly",
        default_kind="Stacked Bar",
        allowed_kinds=["Bar", "Line", "Area", "Stacked Bar", "Stacked Area"],
        x_field="YM_dt",
        y_field="Hours",
        color_field="Tätigkeit",
        x_type="temporal",
        stacked_default=True,
        pie_ok=False,
    )

    cA, cB, cC = st.columns(3)

    with cA:
        _render_chart_block(
            title="Nach Tätigkeit",
            df=type_df,
            key_prefix="viz_type",
            default_kind="Pie",
            allowed_kinds=["Pie", "Donut", "Bar"],
            x_field="Tätigkeit",
            y_field="Hours",
            color_field=None,
            x_type="nominal",
            stacked_default=False,
            pie_ok=True,
        )

    with cB:
        _render_chart_block(
            title="Nach Projekt",
            df=proj_df,
            key_prefix="viz_project",
            default_kind="Bar",
            allowed_kinds=["Bar", "Pie", "Donut"],
            x_field="Projekt",
            y_field="Hours",
            color_field=None,
            x_type="nominal",
            stacked_default=False,
            pie_ok=True,
        )

    with cC:
        _render_chart_block(
            title="Nach Mitarbeiter",
            df=mitarbeiter_df,
            key_prefix="viz_mitarbeiter",
            default_kind="Pie",
            allowed_kinds=["Pie", "Donut", "Bar"],
            x_field="Mitarbeiter",
            y_field="Hours",
            color_field=None,
            x_type="nominal",
            stacked_default=False,
            pie_ok=True,
        )

    st.markdown("### Top Kodierungen (Aufgaben) nach Stunden")
    top_kod = (
        x.assign(Kod=x["Kodierung"].fillna("").astype(str).str.strip())
        .groupby("Kod")["Hours"].sum()
        .sort_values(ascending=False)
        .head(20)
    )
    top_kod = top_kod[top_kod.index != ""]
    if not top_kod.empty:
        st.dataframe(
            top_kod.reset_index().rename(columns={"Kod": "Kodierung", "Hours": "Stunden"}),
            use_container_width=True,
            height=340,
        )
    else:
        st.caption("Keine Kodierungen vorhanden/gefüllt.")

    st.markdown("### Detail (optional)")
    with st.expander("Rohdaten anzeigen"):
        show_cols = ["Mitarbeiter", "Datum_dt", "Projekt", "Tätigkeit", "Hours", "km", "Kodierung", "Info",
                     "Abgerechnet"]
        show_cols = [c for c in show_cols if c in x.columns]
        st.dataframe(x[show_cols].sort_values(["Datum_dt", "Mitarbeiter"], ascending=[False, True]),
                     use_container_width=True, height=420)


def main() -> None:
    if st is None:
        raise RuntimeError("Streamlit ist nicht installiert. Bitte `pip install streamlit` ausführen.")
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    st.caption("Lokale Webapp für Tätigkeiten-Erfassung und Einsatzbericht-Auswertung (Excel-basiert, MVP)")

    remembered_path = _safe_str(st.session_state.get("excel_path", "")).strip()
    default_path = ""
    if remembered_path:
        try:
            remembered_resolved = _resolve_excel_path(remembered_path)
            if remembered_resolved.exists():
                default_path = remembered_path
        except Exception:
            default_path = ""
    if not default_path:
        for cand in _default_excel_candidates():
            if cand.exists():
                try:
                    script_dir = Path(__file__).resolve().parent
                    default_path = str(cand.resolve().relative_to(script_dir))
                except Exception:
                    default_path = str(cand.resolve())
                break
    if not default_path:
        default_path = "data/Tätigkeiten_Überblick.xlsx"
    pending_import_path = _safe_str(st.session_state.get("_pending_import_excel_path", "")).strip()
    if pending_import_path:
        st.session_state["excel_path_input"] = pending_import_path
        st.session_state["excel_path"] = pending_import_path
        st.session_state["_pending_import_excel_path"] = ""
    if _safe_str(st.session_state.get("excel_path_input", "")).strip():
        try:
            _tmp_vis = _resolve_excel_path(_safe_str(st.session_state.get("excel_path_input", "")).strip())
            if not _tmp_vis.exists():
                st.session_state["excel_path_input"] = default_path
        except Exception:
            st.session_state["excel_path_input"] = default_path
    else:
        st.session_state["excel_path_input"] = default_path

    with st.sidebar:
        st.header("Datei")
        excel_path = st.text_input("Pfad zur Excel-Datei", key="excel_path_input")
        uploaded_excel = st.file_uploader(
            "Excel-Datei importieren (.xlsx)",
            type=["xlsx"],
            key="excel_upload_file",
            help="Lädt eine Excel-Datei hoch und speichert sie als lokale Arbeitskopie (für Bearbeitung + Excel-Druck/PDF).",
        )

        if uploaded_excel is not None:
            if st.button("Upload als Arbeitskopie übernehmen", key="import_uploaded_excel_btn"):
                try:
                    imported_path = _store_uploaded_excel(uploaded_excel)
                    st.session_state["_pending_import_excel_path"] = str(imported_path)
                    st.session_state["excel_path"] = str(imported_path)
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Import fehlgeschlagen: {e}")
        st.session_state["excel_path"] = excel_path
        reload_clicked = st.button("Neu laden")
        st.info(
            "Hinweis: Relative Pfade sind erlaubt.\n\n"
            "Unter Windows nutzt die App Microsoft Excel (via pywin32) für den originalgetreuen Druck/PDF-Export. "
            "Unter macOS wird stattdessen 'openpyxl' in Kombination mit AppleScript verwendet, um Microsoft Excel fernzusteuern."
        )

    if reload_clicked:
        st.cache_data.clear()

    try:
        resolved_path = _resolve_excel_path(excel_path)
        if resolved_path.exists():
            mtime = resolved_path.stat().st_mtime
            data = _cached_load_workbook_data(str(resolved_path), mtime)
        else:
            data = load_workbook_data(excel_path)  # triggers FileNotFoundError cleanly
    except Exception as e:
        st.error(f"Datei konnte nicht geladen werden: {e}")
        st.stop()

    df = data.taetigkeiten_df.copy()
    team_df = data.team_df.copy()
    lookups = data.lookups

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["Tätigkeiten", "Einsatzbericht", "Stammdaten / Debug", "Visualisierung", "Team-Daten"])

    with tab1:
        st.subheader("Tätigkeiten erfassen und bearbeiten")

        y_default, m_default = _project_defaults(df)
        projekte_available = sorted(list(dict.fromkeys([p for p in (lookups.get("projekte") or []) if p] + [p for p in
                                                                                                            df.get(
                                                                                                                "Projekt",
                                                                                                                pd.Series(
                                                                                                                    dtype=str)).dropna().astype(
                                                                                                                str).tolist()
                                                                                                            if p])))
        filt_cols = st.columns([1, 1, 2, 1])
        with filt_cols[0]:
            f_year = st.number_input("Jahr-Filter", min_value=2000, max_value=2100, value=int(y_default), step=1)
        with filt_cols[1]:
            f_month = st.selectbox("Monat-Filter", options=list(range(1, 13)),
                                   index=max(0, min(11, int(m_default) - 1)))
        with filt_cols[2]:
            f_project = st.selectbox("Projekt-Filter", options=[""] + projekte_available, index=0)
        with filt_cols[3]:
            f_include_abg = st.checkbox("abgerechnete zeigen", value=True)

        projekte_opts = projekte_available if projekte_available else [""]
        typen_opts = list(dict.fromkeys(lookups.get("taetigkeit_typen") or ["F", "R", "I", "S", "K"]))
        ja_nein_opts = list(dict.fromkeys(lookups.get("ja_nein") or ["ja", "nein"]))

        kod_opts = list(dict.fromkeys(
            [""] + (lookups.get("relevante_kodierungen") or []) + (lookups.get("kodierung_aufgaben") or [])
        ))

        interne_opts = list(dict.fromkeys([""] + (lookups.get("interne_projekte") or [])))

        filtered = _filtered_taetigkeiten(df, int(f_year), int(f_month), f_project, include_abgerechnet=f_include_abg)
        st.markdown("### Tätigkeiten (Inline bearbeiten)")

        if filtered.empty:
            st.info("Keine Tätigkeiten für den aktuellen Filter gefunden.")
        else:
            editor_cols = [
                "_excel_row",
                "Datum",
                "Projekt",
                "Zeit von",
                "Zeit bis",
                "Pause_Min",
                "Zahl",
                "Dauer",
                "km",
                "Tätigkeit",
                "Kodierung",
                "Interne Projekte",
                "Info",
                "Abgerechnet",
                "eingetragen",
            ]

            editor_df = filtered.copy()

            for c in editor_cols:
                if c not in editor_df.columns:
                    editor_df[c] = None

            editor_df["Löschen"] = False

            data_editor_fn = getattr(st, "data_editor", None) or getattr(st, "experimental_data_editor")

            def _calc_dauer_str(zv, zb, pause_min, zahl) -> str:
                h = _compute_hours_decimal(zv, zb, int(pause_min or 0))
                if h is None:
                    try:
                        if zahl is not None and pd.notna(zahl):
                            h = float(zahl)
                        else:
                            return ""
                    except Exception:
                        return ""
                mins = int(round(h * 60))
                return f"{mins // 60:02d}:{mins % 60:02d}"

            editor_df["Dauer"] = editor_df.apply(
                lambda r: _calc_dauer_str(r.get("Zeit von"), r.get("Zeit bis"), r.get("Pause_Min"), r.get("Zahl")),
                axis=1
            )

            edited_df = data_editor_fn(
                editor_df[editor_cols + ["Löschen"]],
                key="taetigkeiten_inline_editor",
                use_container_width=True,
                height=420,
                num_rows="dynamic",
                hide_index=True,
                disabled=["_excel_row", "Dauer"],
                column_config={
                    "_excel_row": st.column_config.NumberColumn("Excel-Zeile",
                                                                help="Technische Zeilen-ID (nicht ändern)"),
                    "Datum": st.column_config.DateColumn("Datum", format="DD.MM.YYYY"),
                    "Projekt": st.column_config.SelectboxColumn("Projekt", options=projekte_opts),
                    "Zeit von": st.column_config.TimeColumn("Zeit von", format="HH:mm"),
                    "Zeit bis": st.column_config.TimeColumn("Zeit bis", format="HH:mm"),
                    "Pause_Min": st.column_config.NumberColumn("Pause (Min)", min_value=0, max_value=600, step=5),
                    "Zahl": st.column_config.NumberColumn("Zeit (h)",
                                                          help="Dezimalstunden für manuelle Angaben (z.B. Organisatorisches)",
                                                          min_value=0.0, step=0.25, format="%.2f"),
                    "Dauer": st.column_config.TextColumn("Dauer",
                                                         help="Berechnet aus Zeit von/bis und Pause (oder Zeit h)",
                                                         width="small"),
                    "km": st.column_config.NumberColumn("km", min_value=0, step=1),
                    "Tätigkeit": st.column_config.SelectboxColumn("Tätigkeit", options=typen_opts),
                    "Kodierung": st.column_config.SelectboxColumn("Kodierung (Aufgabe)", options=kod_opts),
                    "Interne Projekte": st.column_config.SelectboxColumn("Interne Projekte", options=interne_opts),
                    "Info": st.column_config.TextColumn("Leistungsbeschreibung", width="large"),
                    "Abgerechnet": st.column_config.SelectboxColumn("Abgerechnet", options=[""] + ja_nein_opts),
                    "eingetragen": st.column_config.SelectboxColumn("eingetragen", options=[""] + ja_nein_opts),
                    "Löschen": st.column_config.CheckboxColumn("Löschen"),
                },
            )

            st.caption(
                "✔ Du kannst direkt in der Tabelle editieren. "
                "➕ Neue Zeilen unten hinzufügen. "
                "🗑️ Zum Löschen Checkbox 'Löschen' setzen und speichern."
            )

            def _editor_row_to_record(r: pd.Series) -> Dict[str, Any]:
                pause_min = 0
                try:
                    pause_min = int(r.get("Pause_Min") or 0)
                except Exception:
                    pause_min = 0

                km_val = 0
                try:
                    km_val = int(float(r.get("km") or 0) or 0)
                except Exception:
                    km_val = 0

                return {
                    "Datum": _to_date(r.get("Datum")),
                    "Projekt": _safe_str(r.get("Projekt")).strip(),
                    "Zeit von": _to_time(r.get("Zeit von")),
                    "Zeit bis": _to_time(r.get("Zeit bis")),
                    "Pause_Min": pause_min,
                    "Zahl": r.get("Zahl"),
                    "km": km_val,
                    "Tätigkeit": _safe_str(r.get("Tätigkeit")).strip(),
                    "Kodierung": _safe_str(r.get("Kodierung")).strip(),
                    "Interne Projekte": _safe_str(r.get("Interne Projekte")).strip(),
                    "Info": _safe_str(r.get("Info")),
                    "Abgerechnet": _normalize_yes_no(r.get("Abgerechnet")) or _safe_str(r.get("Abgerechnet")).strip(),
                    "eingetragen": _normalize_yes_no(r.get("eingetragen")) or _safe_str(r.get("eingetragen")).strip(),
                }

            def _is_editor_row_blank(r: pd.Series) -> bool:
                return (
                        _is_blank(r.get("Datum")) and
                        _is_blank(r.get("Projekt")) and
                        _is_blank(r.get("Zeit von")) and
                        _is_blank(r.get("Zeit bis")) and
                        _is_blank(r.get("Tätigkeit")) and
                        _is_blank(r.get("Info")) and
                        _is_blank(r.get("Abgerechnet")) and
                        _is_blank(r.get("eingetragen"))
                )

            if st.button("Tabellenänderungen speichern", key="save_inline_table"):
                orig_by_row = {}
                original_excel_rows = set()
                for _, r in editor_df.iterrows():
                    exr = r.get("_excel_row")
                    if exr is None or (isinstance(exr, float) and pd.isna(exr)):
                        continue
                    try:
                        row_idx = int(exr)
                        orig_by_row[row_idx] = _editor_row_to_record(r)
                        original_excel_rows.add(row_idx)
                    except Exception:
                        continue

                updates: List[Tuple[int, Dict[str, Any]]] = []
                inserts: List[Dict[str, Any]] = []
                deletes: List[int] = []
                kept_excel_rows = set()

                for _, r in edited_df.iterrows():
                    exr = r.get("_excel_row")
                    mark_delete = bool(r.get("Löschen", False))
                    is_new = (exr is None) or (isinstance(exr, float) and pd.isna(exr))

                    if is_new:
                        if _is_editor_row_blank(r):
                            continue
                        rec = _editor_row_to_record(r)
                        if rec.get("Projekt") and rec.get("Datum"):
                            inserts.append(rec)
                        continue

                    row_excel = int(exr)
                    kept_excel_rows.add(row_excel)

                    if mark_delete:
                        deletes.append(row_excel)
                        continue

                    new_rec = _editor_row_to_record(r)
                    old_rec = orig_by_row.get(row_excel)

                    if old_rec is None or _key_for_import(new_rec) != _key_for_import(old_rec):
                        updates.append((row_excel, new_rec))

                for r_id in original_excel_rows:
                    if r_id not in kept_excel_rows and r_id not in deletes:
                        deletes.append(r_id)

                def _mutator_inline(wb):
                    ws = wb[TAETIGKEITEN_SHEET]
                    for r in deletes:
                        _clear_taetigkeit_row(ws, r)
                    for r, rec in updates:
                        _write_taetigkeit_row(ws, r, rec)
                    row_idx = _find_next_write_row(ws, key_cols=KEY_COLS_FOR_EMPTY_CHECK)
                    for rec in inserts:
                        _write_taetigkeit_row(ws, row_idx, rec)
                        row_idx += 1

                ok, msg = _save_workbook(data.path, _mutator_inline)
                if ok:
                    st.success(
                        f"Gespeichert. Updates: {len(updates)}, Neu: {len(inserts)}, Gelöscht: {len(deletes)}. {msg}"
                    )
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(msg)

        st.markdown("---")
        c1, c2 = st.columns(2)

        with c1:
            st.markdown("### Neue Tätigkeit anlegen")
            with st.form("add_form", clear_on_submit=False):
                defaults_add = {
                    "Datum": dt.date.today(),
                    "Projekt": f_project or (projekte_available[0] if projekte_available else ""),
                    "Tätigkeit": "F" if "F" in (lookups.get("taetigkeit_typen") or []) else
                    (lookups.get("taetigkeit_typen") or [""])[0],
                    "Abgerechnet": "nein" if "nein" in (lookups.get("ja_nein") or []) else "",
                }
                rec_add = _render_taetigkeit_form("add", lookups, defaults=defaults_add)
                submit_add = st.form_submit_button("Eintrag speichern")
            if submit_add:
                def _mutator_add(wb):
                    ws = wb[TAETIGKEITEN_SHEET]
                    row_idx = _find_next_write_row(ws, key_cols=KEY_COLS_FOR_EMPTY_CHECK)
                    _write_taetigkeit_row(ws, row_idx, rec_add)

                ok, msg = _save_workbook(data.path, _mutator_add)
                if ok:
                    st.success(msg)
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(msg)

        with c2:
            st.markdown("### Bestehenden Eintrag bearbeiten")
            if filtered.empty:
                st.info("Zum Bearbeiten zuerst oben einen Filter mit Treffern wählen.")
            else:
                options = filtered.to_dict(orient="records")
                labels = [_display_row_label(pd.Series(o)) for o in options]
                idx = st.selectbox("Eintrag auswählen", options=list(range(len(options))),
                                   format_func=lambda i: labels[i], key="edit_row_selector")
                selected = options[int(idx)]

                with st.form("edit_form", clear_on_submit=False):
                    rec_edit = _render_taetigkeit_form("edit", lookups, defaults=selected)
                    save_edit = st.form_submit_button("Änderungen speichern")
                    delete_row = st.form_submit_button("Zeile leeren (vorsichtig)")

                row_excel = int(selected["_excel_row"])
                if save_edit:
                    def _mutator_edit(wb):
                        ws = wb[TAETIGKEITEN_SHEET]
                        _write_taetigkeit_row(ws, row_excel, rec_edit)

                    ok, msg = _save_workbook(data.path, _mutator_edit)
                    if ok:
                        st.success(f"Zeile {row_excel} aktualisiert. {msg}")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(msg)
                if delete_row:
                    def _mutator_delete(wb):
                        ws = wb[TAETIGKEITEN_SHEET]
                        _clear_taetigkeit_row(ws, row_excel)

                    ok, msg = _save_workbook(data.path, _mutator_delete)
                    if ok:
                        st.success(f"Zeile {row_excel} geleert. {msg}")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(msg)

        st.markdown("---")
        st.markdown("### Schnellaktion")
        if not filtered.empty:
            if st.button("Gefilterte Treffer als 'abgerechnet = ja' markieren"):
                rows_to_mark = [int(r) for r in filtered["_excel_row"].tolist()]

                def _mutator_mark(wb):
                    ws = wb[TAETIGKEITEN_SHEET]
                    for r in rows_to_mark:
                        ws.cell(r, 13).value = "ja"

                ok, msg = _save_workbook(data.path, _mutator_mark)
                if ok:
                    st.success(f"{len(rows_to_mark)} Einträge markiert. {msg}")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(msg)
        else:
            st.caption("Keine Treffer für Schnellaktion.")

        st.markdown("---")
        st.markdown("### Migration: Einsatzbericht importieren")

        with st.expander("Einsatzbericht(e) (Excel-Layout) hochladen und importieren"):

            import_mode = st.radio("Wohin sollen die Dateien importiert werden?", [
                "In Meine Tätigkeiten (Mein Journal)",
                "In Team-Tätigkeiten (Fürs Projekt-Controlling)"
            ],
                                   help="Wähle 'Team-Tätigkeiten', um die Berichte anderer Mitarbeiter strikt von deinen eigenen EBs zu trennen.")

            report_files = st.file_uploader(
                "Einsatzbericht-Excel auswählen (.xlsx)",
                type=["xlsx"],
                accept_multiple_files=True,
                key="upload_reports",
            )

            default_km = st.number_input("km (Default für importierte Zeilen)", min_value=0, value=0, step=1,
                                         key="import_km_default")
            set_eingetragen = st.checkbox("eingetragen automatisch = ja", value=True, key="import_set_eingetragen")
            set_abgerechnet = st.checkbox("abgerechnet automatisch = nein", value=True, key="import_set_abgerechnet")

            if report_files:
                rev_kod_map = _build_reverse_kod_map_eb_to_aufgabe(lookups)
                is_team_mode = "Team" in import_mode
                existing_keys = _existing_keys_for_master(team_df if is_team_mode else df, team=is_team_mode)
                seen_source_rows = set()

                all_records: List[Dict[str, Any]] = []
                meta_list = []

                target_projects = sorted(list(dict.fromkeys(
                    [p for p in (lookups.get("projekte") or []) if p] +
                    [p for p in df.get("Projekt", pd.Series(dtype=str)).dropna().astype(str).tolist() if p]
                )))

                for i, uf in enumerate(report_files, start=1):
                    try:
                        file_name = getattr(uf, "name", f"report_{i}.xlsx")

                        allowed_arts_tuple = tuple(lookups.get("taetigkeit_typen") or ["F", "R", "I", "S", "K"])
                        meta, lines = _parse_and_store_uploaded_report(file_name, uf.getvalue(), allowed_arts_tuple)

                        st.write("---")
                        st.write(f"**Datei {i}:** `{file_name}`")
                        st.write(f"Zeilen erkannt: {len(lines)}")
                        meta_list.append(meta)

                        detected_project = meta.get("project") or ""

                        rc1, rc2 = st.columns(2)
                        with rc1:
                            st.write(f"Projekt/Firma (aus Excel): `{detected_project or '-'}`")
                            st.write(f"Monat/Jahr: {meta.get('month')}/{meta.get('year')}")

                        guessed_proj = _guess_project(meta, file_name, lookups, target_projects)
                        default_target = guessed_proj if guessed_proj in target_projects else (
                            target_projects[0] if target_projects else ""
                        )

                        with rc2:
                            target_project = st.selectbox(
                                f"Ziel-Projekt festlegen",
                                options=target_projects if target_projects else [""],
                                index=(
                                    target_projects.index(default_target) if default_target in target_projects else 0),
                                key=f"import_report_target_{i}",
                            )
                            mitarbeiter_name = ""
                            if is_team_mode:
                                mitarbeiter_name = st.text_input(f"Mitarbeiter-Name festlegen",
                                                                 value=meta.get("name", f"Kollege {i}"),
                                                                 key=f"import_team_name_{i}")

                        if lines.empty:
                            st.info("Keine gültigen Tätigkeits-Zeilen im Einsatzbericht gefunden.")
                            continue

                        for _, row in lines.iterrows():
                            datum = row.get("Datum")
                            beginn = row.get("Beginn")
                            ende = row.get("Ende")
                            pause_min = int(row.get("Pause_Min") or 0)
                            art = _safe_str(row.get("Art")).strip()
                            kod_eb = _safe_str(row.get("Kodierung_EB")).strip()
                            text = _safe_str(row.get("Leistungsbeschreibung")).strip()

                            aufgabe = rev_kod_map.get(kod_eb.lower(), "")
                            info = text
                            if kod_eb and not aufgabe:
                                all_aufgaben = [a.lower() for a in (lookups.get("kodierung_aufgaben") or [])]
                                if kod_eb.lower() in all_aufgaben:
                                    idx = all_aufgaben.index(kod_eb.lower())
                                    aufgabe = (lookups.get("kodierung_aufgaben") or [])[idx]
                                else:
                                    aufgabe = kod_eb

                            rec = {
                                "Datum": datum,
                                "Projekt": target_project,
                                "Zeit von": beginn,
                                "Zeit bis": ende,
                                "Pause_Min": pause_min,
                                "km": int(default_km),
                                "Tätigkeit": art,
                                "Kodierung": aufgabe or "",
                                "Interne Projekte": "",
                                "Info": info,
                                "Abgerechnet": "nein" if set_abgerechnet else "",
                                "eingetragen": "ja" if set_eingetragen else "",
                                "Zahl": row.get("Zeit_h"),
                            }
                            if is_team_mode:
                                rec["Mitarbeiter"] = mitarbeiter_name

                            k = _key_for_import(rec)
                            if is_team_mode:
                                key_tuple = (_safe_str(mitarbeiter_name).strip(),) + k
                                if key_tuple in existing_keys:
                                    continue
                            else:
                                if k in existing_keys:
                                    continue

                            source_row_key = (
                                _safe_str(file_name).strip(),
                                int(row.get("_source_row") or 0),
                                _safe_str(mitarbeiter_name).strip() if is_team_mode else "",
                                target_project,
                            )
                            if source_row_key in seen_source_rows:
                                continue
                            seen_source_rows.add(source_row_key)

                            all_records.append(rec)

                    except Exception as e:
                        st.error(f"Fehler beim Lesen der Datei {i}: {e}")

                st.write("---")
                st.metric("Gesamt neu zu importierende Zeilen", len(all_records))

                if all_records:
                    preview_data = []
                    for r in all_records:
                        d = {
                            "Datum": _format_date(r["Datum"]),
                            "Projekt": r["Projekt"],
                            "Zeit von": _format_time(r["Zeit von"]),
                            "Zeit bis": _format_time(r["Zeit bis"]),
                            "Pause (Min)": r["Pause_Min"],
                            "Zeit (h)": r.get("Zahl"),
                            "Typ": r["Tätigkeit"],
                            "Kodierung (Aufgabe)": r["Kodierung"],
                            "Info": _safe_str(r["Info"])[:80],
                        }
                        if is_team_mode:
                            d = {"Mitarbeiter": r.get("Mitarbeiter")} | d  # Prepend Mitarbeiter
                        preview_data.append(d)

                    preview = pd.DataFrame(preview_data)
                    st.dataframe(preview, use_container_width=True, height=260)

                    if st.button("Import durchführen (Schreiben)", key="commit_report_import"):
                        def _mutator_import_reports(wb):
                            if is_team_mode:
                                ws = _ensure_team_sheet(wb)
                                row_idx = _find_next_write_row(ws, key_cols=TEAM_KEY_COLS_FOR_EMPTY_CHECK)
                                for rec in all_records:
                                    _write_team_row(ws, row_idx, rec)
                                    row_idx += 1
                            else:
                                ws = wb[TAETIGKEITEN_SHEET]
                                row_idx = _find_next_write_row(ws, key_cols=KEY_COLS_FOR_EMPTY_CHECK)
                                for rec in all_records:
                                    _write_taetigkeit_row(ws, row_idx, rec)
                                    row_idx += 1

                        ok, msg = _save_workbook(data.path, _mutator_import_reports)
                        if ok:
                            st.success(f"Import OK: {len(all_records)} Zeilen übernommen. {msg}")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(msg)

    with tab2:
        st.subheader("Einsatzbericht (Web-Ansicht)")

        rep_col1, rep_col2, rep_col3, rep_col4 = st.columns([1, 1, 2, 1])
        rep_projects = sorted(list(dict.fromkeys([p for p in (lookups.get("projekte") or []) if p] + [p for p in
                                                                                                      df.get("Projekt",
                                                                                                             pd.Series(
                                                                                                                 dtype=str)).dropna().astype(
                                                                                                          str).tolist()
                                                                                                      if p])))
        y_default, m_default = _project_defaults(df)
        with rep_col1:
            r_year = st.number_input("Jahr", min_value=2000, max_value=2100, value=int(y_default), step=1,
                                     key="rep_year")
        with rep_col2:
            r_month = st.selectbox("Monat", options=list(range(1, 13)), index=max(0, min(11, int(m_default) - 1)),
                                   key="rep_month")
        with rep_col3:
            default_project = "ABS" if "ABS" in rep_projects else (rep_projects[0] if rep_projects else "")
            r_project = st.selectbox("Projekt", options=rep_projects if rep_projects else [""], index=(
                rep_projects.index(default_project) if default_project in rep_projects else 0), key="rep_project")
        with rep_col4:
            include_abgerechnet = st.checkbox("abgerechnete einschließen", value=False)

        report_df = _build_report(df, lookups, int(r_year), int(r_month), r_project, include_abgerechnet)
        sums = _summaries_from_report(report_df)

        info = lookups.get("projekt_infos", {}).get(r_project, {})
        p1, p2 = st.columns(2)
        with p1:
            st.markdown("#### Leistungsnehmer")
            st.write(f"**Firma:** {_safe_str(info.get('Kunde')) or '-'}")
            st.write(f"**Straße:** {_safe_str(info.get('Straße')) or '-'}")
            st.write(f"**Ort:** {_safe_str(info.get('Ort')) or '-'}")
            st.write(f"**Kontakt:** {_safe_str(info.get('Ansprechpartner')) or '-'}")
        with p2:
            st.markdown("#### Report-Kontext")
            st.write(f"**Projekt:** {r_project}")
            st.write(f"**Monat/Jahr:** {int(r_month):02d}/{int(r_year)}")
            st.write(f"**Projektadresse (Standard):** {_safe_str(info.get('Projektadresse Standard')) or '-'}")
            alt = _safe_str(info.get('Projektadresse Alternativ'))
            if alt:
                st.write(f"**Projektadresse (Alternativ):** {alt}")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Beratung (F)", f"{sums['F']:.2f} h")
        m2.metric("Reisezeit (R)", f"{sums['R']:.2f} h")
        m3.metric("Kulanz (K)", f"{sums['K']:.2f} h")
        m4.metric("Gesamt", f"{sums['gesamt']:.2f} h")

        if len(report_df) > REPORT_ROWS_PER_PAGE:
            st.info(
                f"Es werden automatisch mehrere Formularseiten erzeugt: {len(report_df)} Positionen = {math.ceil(len(report_df) / REPORT_ROWS_PER_PAGE)} Seiten à {REPORT_ROWS_PER_PAGE} Positionen.")

        st.markdown("### Original-Einsatzbericht (Excel-Layout)")
        st.caption(
            "Unter Windows nutzt die App Microsofts COM-System für den originalgetreuen PDF-Export. "
            "Unter macOS nutzt sie eine universelle Fallback-Methode in Kombination mit AppleScript."
        )

        export_dir = data.path.parent / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        base_export_name = f"Einsatzbericht_{r_project}_{int(r_year)}-{int(r_month):02d}"
        default_pdf_name = f"{base_export_name}.pdf"
        default_xlsx_name = f"{base_export_name}.xlsx"

        # Dynamischer Key: Wenn sich Projekt/Datum ändern, baut sich das Textfeld komplett mit dem neuen Wert auf!
        pdf_rel_or_abs = st.text_input(
            "PDF-Ausgabepfad",
            value=str((export_dir / default_pdf_name)),
            help="Relativer oder absoluter Pfad. (Aktualisiert sich automatisch bei Projekt-/Datumswechsel)",
            key=f"pdf_path_{r_project}_{r_year}_{r_month}",
        )

        xlsx_rel_or_abs = st.text_input(
            "Excel-Kopie-Ausgabepfad",
            value=str((export_dir / default_xlsx_name)),
            help="Speichert eine vorbereitete Kopie der Excel.",
            key=f"xlsx_path_{r_project}_{r_year}_{r_month}",
        )

        b1, b2, b3, b4 = st.columns(4)

        with b1:
            if st.button("Original in Excel öffnen", key="open_original_excel"):
                ok, msg, _ = _excel_original_report_action(
                    data.path, int(r_year), int(r_month), r_project, action="open", report_df=report_df
                )
                (st.success if ok else st.error)(msg)

        with b2:
            if st.button("Als PDF generieren", key="export_original_pdf"):
                pdf_target = Path(pdf_rel_or_abs).expanduser()
                if not pdf_target.is_absolute():
                    pdf_target = (Path.cwd() / pdf_target).resolve()

                ok, msg, exported_files = _excel_original_report_action(
                    data.path, int(r_year), int(r_month), r_project, action="pdf", pdf_output_path=pdf_target,
                    report_df=report_df
                )
                if ok:
                    st.success(msg)
                    st.session_state["pdf_files_ready"] = exported_files
                else:
                    st.error(msg)

            # Download-Button außerhalb der if-Abfrage rendern
            if st.session_state.get("pdf_files_ready"):
                for p in st.session_state["pdf_files_ready"]:
                    if Path(p).exists():
                        st.download_button(
                            label=f"PDF Downloaden: {Path(p).name}",
                            data=Path(p).read_bytes(),
                            file_name=Path(p).name,
                            mime="application/pdf",
                            key=f"dl_pdf_btn_{Path(p).name}"
                        )

        with b3:
            if st.button("Als Excel-Kopie generieren", key="export_original_xlsx_copy"):
                xlsx_target = Path(xlsx_rel_or_abs).expanduser()
                if not xlsx_target.is_absolute():
                    xlsx_target = (Path.cwd() / xlsx_target).resolve()

                ok, msg, exported_files = _excel_original_report_action(
                    data.path, int(r_year), int(r_month), r_project, action="xlsx", xlsx_output_path=xlsx_target,
                    report_df=report_df
                )
                if ok:
                    st.success(msg)
                    st.session_state["xlsx_files_ready"] = exported_files
                else:
                    st.error(msg)

            # Download-Button außerhalb der if-Abfrage rendern
            if st.session_state.get("xlsx_files_ready"):
                for p in st.session_state["xlsx_files_ready"]:
                    if Path(p).exists():
                        st.download_button(
                            label=f"Excel Downloaden: {Path(p).name}",
                            data=Path(p).read_bytes(),
                            file_name=Path(p).name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_xlsx_btn_{Path(p).name}"
                        )

        with b4:
            if st.button("Original direkt drucken", key="print_original_excel"):
                ok, msg, _ = _excel_original_report_action(
                    data.path, int(r_year), int(r_month), r_project, action="print", report_df=report_df
                )
                (st.success if ok else st.error)(msg)

        if report_df.empty:
            st.warning("Keine Einträge für diesen Einsatzbericht gefunden.")
            st.caption(
                "Tipp: 'abgerechnete einschließen' aktivieren, falls die Beispiel-Datei nur bereits markierte Einträge enthält.")
        else:
            st.dataframe(report_df.drop(columns=["_excel_row"]), use_container_width=True, height=360)
            csv_bytes = report_df.drop(columns=["_excel_row"]).to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "Tabelle als CSV exportieren",
                data=csv_bytes,
                file_name=f"Einsatzbericht_{r_project}_{r_year}-{int(r_month):02d}.csv",
                mime="text/csv",
                key="dl_csv_export"
            )
            with st.expander("Treffer als abgerechnet markieren"):
                st.caption("Schreibt 'ja' in Spalte 'Abgerechnet' für die aktuell im Report enthaltenen Tätigkeiten.")
                if st.button("Report-Treffer markieren", key="mark_report_done"):
                    rows_to_mark = [int(r) for r in report_df["_excel_row"].tolist()]

                    def _mutator_mark_report(wb):
                        ws = wb[TAETIGKEITEN_SHEET]
                        for r in rows_to_mark:
                            ws.cell(r, 13).value = "ja"

                    ok, msg = _save_workbook(data.path, _mutator_mark_report)
                    if ok:
                        st.success(f"{len(rows_to_mark)} Einträge markiert. {msg}")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(msg)

    with tab3:
        st.subheader("Stammdaten / Debug")
        st.write(f"**Datei:** `{data.path}`")
        st.write(f"**Tätigkeiten (erkannte Datensätze):** {len(df)}")

        c1, c2, c3 = st.columns(3)
        c1.write("**Projekte**")
        c1.dataframe(pd.DataFrame({"Projekt": lookups.get("projekte", [])}), use_container_width=True, height=200)
        c2.write("**Tätigkeitstypen**")
        c2.dataframe(pd.DataFrame({"Typ": lookups.get("taetigkeit_typen", [])}), use_container_width=True, height=200)
        c3.write("**Relevante Kodierungen (aktiv)**")
        c3.dataframe(pd.DataFrame({"Kodierung": lookups.get("relevante_kodierungen", [])}), use_container_width=True,
                     height=200)

        st.markdown("---")
        st.markdown("### Relevante Kodierungen pflegen")
        all_kod_aufgaben = lookups.get("kodierung_aufgaben", []) or []
        if all_kod_aufgaben:
            with st.form("relevante_kodierungen_form", clear_on_submit=False):
                selected_relevante = st.multiselect(
                    "Wähle die abrechnungs-/auswahltrelevanten Kodierungen (Aufgabe)",
                    options=all_kod_aufgaben,
                    default=[k for k in (lookups.get("relevante_kodierungen", []) or []) if k in all_kod_aufgaben],
                    help="Speichert die Auswahl als 'x' in 'Kodierung Joyson' Spalte A.",
                )
                save_relevante = st.form_submit_button("Relevante Kodierungen speichern")
            if save_relevante:
                def _mutator_relevante(wb):
                    ok_inner, msg_inner = _set_relevante_kodierungen(wb, selected_relevante)
                    if not ok_inner:
                        raise ValueError(msg_inner)

                ok, msg = _save_workbook(data.path, _mutator_relevante)
                if ok:
                    st.success(msg)
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(msg)
        else:
            st.info("Keine Kodierungen aus 'Kodierung Joyson' erkannt.")

        st.markdown("---")
        st.markdown("### Projekt-Stammdaten pflegen")
        projekt_infos = lookups.get("projekt_infos", {}) or {}
        projekte_sortiert = sorted(projekt_infos.keys())
        proj_select = st.selectbox(
            "Projekt auswählen",
            options=["<Neues Projekt>"] + projekte_sortiert,
            index=0,
            key="proj_master_select",
        )

        if proj_select == "<Neues Projekt>":
            proj_defaults = {
                "Projekt": "",
                "Kunde": "",
                "Straße": "",
                "Ort": "",
                "Ansprechpartner": "",
                "Projektadresse Standard": "",
                "Projektadresse Alternativ": "",
            }
            original_project_name = ""
        else:
            proj_defaults = dict(projekt_infos.get(proj_select, {}))
            original_project_name = proj_select

        with st.form("projekt_stammdaten_form", clear_on_submit=False):
            pc1, pc2 = st.columns(2)
            with pc1:
                p_name = st.text_input("Projekt (Kürzel/Name)", value=_safe_str(proj_defaults.get("Projekt")))
                p_kunde = st.text_input("Kunde", value=_safe_str(proj_defaults.get("Kunde")))
                p_ansp = st.text_input("Ansprechpartner", value=_safe_str(proj_defaults.get("Ansprechpartner")))
                p_addr_std = st.text_input("Projektadresse Standard",
                                           value=_safe_str(proj_defaults.get("Projektadresse Standard")))
            with pc2:
                p_str = st.text_input("Straße", value=_safe_str(proj_defaults.get("Straße")))
                p_ort = st.text_input("Ort", value=_safe_str(proj_defaults.get("Ort")))
                p_addr_alt = st.text_input("Projektadresse Alternativ",
                                           value=_safe_str(proj_defaults.get("Projektadresse Alternativ")))
                rename_taet = st.checkbox(
                    "Bei Projekt-Umbenennung auch Tätigkeiten aktualisieren",
                    value=False,
                    disabled=(not original_project_name),
                    help="Ändert Spalte 'Projekt' in 'Tätigkeiten' von altem auf neuen Projektnamen.",
                )

            save_proj = st.form_submit_button("Projekt-Stammdaten speichern")

        if save_proj:
            payload = {
                "Projekt": p_name,
                "Kunde": p_kunde,
                "Straße": p_str,
                "Ort": p_ort,
                "Ansprechpartner": p_ansp,
                "Projektadresse Standard": p_addr_std,
                "Projektadresse Alternativ": p_addr_alt,
            }

            def _mutator_proj(wb):
                ok_inner, msg_inner = _upsert_project_stammdaten(
                    wb,
                    payload,
                    original_project=original_project_name or None,
                    rename_taetigkeiten=bool(rename_taet),
                )
                if not ok_inner:
                    raise ValueError(msg_inner)

            ok, msg = _save_workbook(data.path, _mutator_proj)
            if ok:
                st.success(msg)
                st.cache_data.clear()
                st.rerun()
            else:
                st.error(msg)

        with st.expander("Projekt-Stammdaten (Tabelle)"):
            proj_df = pd.DataFrame.from_dict(lookups.get("projekt_infos", {}), orient="index").reset_index(drop=True)
            st.dataframe(proj_df, use_container_width=True, height=300)

        with st.expander("Kodierung-Mapping (Aufgabe -> Kodierung EB)"):
            km = lookups.get("kodierung_map_eb", {})
            km_df = pd.DataFrame([{"Aufgabe": k, "Kodierung EB": v} for k, v in km.items()])
            st.dataframe(km_df, use_container_width=True, height=300)

    with tab4:
        _render_visualisierung_tab(df, team_df, lookups, data.path, data.milestones_df)

    with tab5:
        st.subheader("Team-Tätigkeiten (Fremddaten)")
        st.write(
            "Diese Daten stammen aus dem Import von Einsatzberichten anderer Mitarbeiter (Modus: 'In Team-Tätigkeiten'). Sie werden für das Budget-Controlling genutzt, aber von deinem eigenen Einsatzbericht (Tab: 'Einsatzbericht') völlig ignoriert.")

        if team_df.empty:
            st.info("Noch keine Team-Daten vorhanden.")
        else:
            show_cols = ["_excel_row", "Mitarbeiter", "Datum", "Projekt", "Zeit von", "Zeit bis", "Pause_Min", "Zahl",
                         "Tätigkeit", "Kodierung", "Info", "Abgerechnet"]
            disp_df = team_df[[c for c in show_cols if c in team_df.columns]].copy()
            if "Datum" in disp_df.columns:
                disp_df["Datum"] = disp_df["Datum"].apply(_format_date)
            if "Zeit von" in disp_df.columns:
                disp_df["Zeit von"] = disp_df["Zeit von"].apply(_format_time)
            if "Zeit bis" in disp_df.columns:
                disp_df["Zeit bis"] = disp_df["Zeit bis"].apply(_format_time)

            st.dataframe(disp_df.sort_values(["Datum", "Mitarbeiter"], ascending=[False, True]),
                         use_container_width=True, height=400)

            with st.expander("Gefahr: Team-Daten komplett leeren"):
                st.warning(
                    "Löscht alle aktuell importierten Team-Tätigkeiten aus der Excel-Tabelle 'Team_Tätigkeiten'.")
                if st.button("Alle Team-Daten verwerfen (Löschen)"):
                    def _clear_all_team(wb):
                        if TEAM_SHEET in wb.sheetnames:
                            ws = wb[TEAM_SHEET]
                            # Start ab Zeile 2 bis Ende
                            for r in range(2, ws.max_row + 1):
                                for c in range(1, 16):
                                    ws.cell(r, c).value = None

                    ok, msg = _save_workbook(data.path, _clear_all_team)
                    if ok:
                        st.success("Tabelle 'Team_Tätigkeiten' wurde erfolgreich geleert.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"Fehler beim Löschen: {msg}")


if __name__ == "__main__":
    main()
